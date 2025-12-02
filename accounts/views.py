from multiprocessing import context
from django.utils import timezone
from django.contrib.auth import authenticate, login
from django.contrib.auth.forms import AuthenticationForm
from django.contrib.auth.models import User
from django.utils.timezone import now
from networkx import center
from accounts.utils import QUARTER_CHOICES, get_month_range
from .forms import  CustomUserCreationForm, DayRevisionFeedbackForm, DayTrainingReportForm, DepartmentCoordinatorRegistrationForm,  DocumentCommentForm, DocumentDayFileForm, DocumentDayForm, ExtensionistRegistrationForm, LinkageForm, MOAResourceForm, RevisionFeedbackForm
from django.http import Http404, HttpResponse, HttpResponseForbidden
from .models import  AccountType, CompletionRevisionFeedback, DayRevisionFeedback, Department,  DocumentComment, DocumentDay, DocumentDayFile, DocumentFile, Document, DocumentRevisionFeedback, Linkage, MOAResource
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect, get_object_or_404
from .forms import DocumentUploadForm, CompletionUploadForm
from django.contrib import messages
from django.contrib.auth import get_user_model
from collections import defaultdict
from .models import ChatMessage
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.core import serializers
from django.utils.timezone import localtime
from django.http import HttpResponseForbidden
from django.views.decorators.http import require_POST
from .models import CustomUser
from accounts import models  # or whatever your custom user model is named
from django.db.models import Q 
from datetime import datetime
from calendar import month_name
from django.shortcuts import render, redirect, get_object_or_404
from .models import AnnualReport, ExtensionActivity
from .forms import AnnualReportForm, ExtensionActivityForm
from django.http import FileResponse
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
import io
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from .models import QuarterlyReport, ExtensionProject
from .forms import QuarterlyReportForm, ExtensionProjectForm
from reportlab.lib.pagesizes import landscape, letter
from reportlab.platypus import Table, TableStyle, PageBreak, Paragraph, SimpleDocTemplate
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import Spacer
from reportlab.lib.units import inch
import json
from django.contrib.auth import logout
from accounts.models import AccountType


def base(request):
    return render(request, 'accounts/base.html')

def video_view(request):
    image_list = ['6.jpg', '7.jpg', '8.jpg', '9.jpg', '10.jpg']
    return render(request, 'accounts/video.html', {'images': image_list})

def photos_view(request):
    return render(request, 'accounts/photos.html')

def objectives_view(request):
    return render(request, 'accounts/objectives.html')

def moa_view(request):
    return render(request, 'accounts/moa.html')

from django.contrib.auth import login
from django.contrib.auth.forms import AuthenticationForm
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect
from .models import AccountType


def login_view(request):
    if request.method == "POST":
        form = AuthenticationForm(request, data=request.POST)
        if form.is_valid():
            user = form.get_user()
            login(request, user)

            # Redirect based on account type
            if user.account_type == AccountType.SUPER_ADMIN:
                return redirect('super_admin_dashboard')
            elif user.account_type == AccountType.CAMPUS_ADMIN:
                return redirect('campus_admin_dashboard')
            elif user.account_type == AccountType.STAFF_EXTENSIONIST:
                return redirect('staff_extensionist_dashboard')
            elif user.account_type == AccountType.DEPARTMENT_COORDINATOR:
                return redirect('department_coordinator_dashboard')
            elif user.account_type == AccountType.EXTENSIONIST:
                return redirect('extensionist_dashboard')
            else:
                return redirect('permission_denied')
    else:
        form = AuthenticationForm()

    return render(request, 'accounts/login.html', {'form': form})


def logout_view(request):
    logout(request)  # This clears the session
    return redirect('home2')
# Registration View
@login_required
def register(request):
    if request.method == 'POST':
        form = CustomUserCreationForm(request.POST, request.FILES)  # ✅ Include request.FILES
        if form.is_valid():
            user = form.save()
            login(request, user)
            return redirect('campus_admin_dashboard')
    else:
        form = CustomUserCreationForm()
    return render(request, 'accounts/register.html', {'form': form})

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect
from .forms import DepartmentCoordinatorRegistrationForm

@login_required
def register_department_coordinator(request):
    if request.method == 'POST':
        form = DepartmentCoordinatorRegistrationForm(request.POST, request.FILES)
        if form.is_valid():
            user = form.save(commit=False)
            user.account_type = 'Department Coordinator'
            user.save()
            # ✅ Only success message relevant to this page
            messages.success(request, 'Department Coordinator account has been successfully created.')
            return redirect('register_department_coordinator')
        else:
            # ✅ Show only error messages for form validation
            messages.error(request, 'There was an error creating the account. Please check the form and try again.')
    else:
        form = DepartmentCoordinatorRegistrationForm()

    # ✅ Filter messages to only include error-type messages for this page
    storage = messages.get_messages(request)
    filtered_messages = [m for m in storage if 'error' in m.tags or 'danger' in m.tags]

    context = {
        'form': form,
        'messages': filtered_messages,  # Only pass error messages to template
    }
    return render(request, 'accounts/register_department_coordinator.html', context)


from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect
from .forms import ExtensionistRegistrationForm

@login_required
def register_extensionist(request):
    if request.method == 'POST':
        form = ExtensionistRegistrationForm(request.POST)
        if form.is_valid():
            user = form.save(commit=False)
            user.account_type = 'Extensionist'
            
            # ✅ Automatically set department from coordinator's department
            if hasattr(request.user, 'department') and request.user.department:
                user.department = request.user.department
            
            user.save()
            # ✅ Success message for this specific form
            messages.success(request, 'Extensionist account has been successfully created.')
            return redirect('register_extensionist')
        else:
            # ✅ Error message for failed registration
            messages.error(request, 'There was an error creating the account. Please check the form and try again.')
    else:
        form = ExtensionistRegistrationForm()
        
        # ✅ Pre-fill department field with coordinator's department and make it read-only
        if hasattr(request.user, 'department') and request.user.department:
            form.fields['department'].initial = request.user.department
            form.fields['department'].widget.attrs['readonly'] = True
            form.fields['department'].disabled = True

    # ✅ Only keep error-type messages for rendering (hide unrelated ones)
    storage = messages.get_messages(request)
    filtered_messages = [m for m in storage if 'error' in m.tags or 'danger' in m.tags]

    context = {
        'form': form,
        'messages': filtered_messages,  # Pass only error messages to the template
    }
    return render(request, 'accounts/register_extensionist.html', context)


User = get_user_model()

from django.contrib.auth.decorators import login_required
from django.contrib.auth.forms import PasswordChangeForm
from django.contrib.auth import update_session_auth_hash
from django.shortcuts import render, redirect
from .forms import UserEditForm

@login_required
def update_account(request):
    user_to_edit = request.user  # No need for user_id

    if request.method == 'POST':
        user_form = UserEditForm(request.POST, request.FILES, instance=user_to_edit)
        if user_form.is_valid():
            user_form.save()
            messages.success(request, 'Account updated successfully.')
            return redirect('update_account')
    else:
        user_form = UserEditForm(instance=user_to_edit)

    password_form = PasswordChangeForm(user_to_edit)

    return render(request, 'accounts/update_account.html', {
        'user_form': user_form,
        'password_form': password_form,
    })


@login_required
def edit_user_view(request, user_id):
    user_to_edit = get_object_or_404(CustomUser, id=user_id)

    # Restrict access
    is_self = request.user.id == user_to_edit.id
    is_allowed_editor = request.user.account_type in ['Staff Extensionist', 'Department Coordinator']

    if not is_self and not is_allowed_editor:
        return HttpResponseForbidden("You are not allowed to edit this account.")

    if request.method == 'POST':
        user_form = UserEditForm(request.POST, request.FILES, instance=user_to_edit)
        if user_form.is_valid():
            user_form.save()
            return redirect('user_hierarchy')
    else:
        user_form = UserEditForm(instance=user_to_edit)

    password_form = None
    if is_self:
        password_form = PasswordChangeForm(request.user)

    return render(request, 'accounts/update_account.html', {
        'user_form': user_form,
        'password_form': password_form,
    })

@login_required
def change_password(request):
    if request.method == 'POST':
        password_form = PasswordChangeForm(request.user, request.POST)
        if password_form.is_valid():
            user = password_form.save()
            update_session_auth_hash(request, user)  # Important!
            return redirect('update_account')
    else:
        password_form = PasswordChangeForm(request.user)

    user_form = UserEditForm(instance=request.user)

    return render(request, 'accounts/update_account.html', {
        'user_form': user_form,
        'password_form': password_form,
    })
@login_required
def delete_extensionist(request, user_id):
    if request.method == 'POST':
        user_to_delete = get_object_or_404(User, id=user_id)
        
        # Add your permission logic here, e.g. only allow certain users to delete
        if request.user.account_type not in ['Campus Extension Coordinator', 'Staff Extensionist']:
            messages.error(request, "You don't have permission to delete users.")
            return redirect('extensionists_list')  # Change to your listing view name
        
        user_to_delete.delete()
        messages.success(request, f"{user_to_delete.full_name} has been deleted.")
        return redirect('extensionists_list')  # Change to your listing view name
    
    # If GET request or others, just redirect
    return redirect('extensionists_list')
@login_required
def user_hierarchy_view(request):
    context = {
        'super_admins': CustomUser.objects.filter(account_type='Super Admin'),
        'campus_admins': CustomUser.objects.filter(account_type='Campus Admin'),
        'staff_extensionists': CustomUser.objects.filter(account_type='Staff Extensionist'),
        'department_coordinators': CustomUser.objects.filter(account_type='Department Coordinator'),
    }
    return render(request, 'accounts/user_hierarchy.html', context)
@login_required
def delete_user(request, user_id):
    allowed_roles = ['Campus Dean', 'Campus Extension Coordinator', 'Staff Extensionist']
    
    # Debug log to check the user's role
    print(f"Current user's role: {request.user.get_account_type_display()}")

    if request.user.get_account_type_display() not in allowed_roles:
        raise Http404("You are not authorized to delete users.")
    
    user = get_object_or_404(CustomUser, id=user_id)

    # Deleting the user
    user.delete()

    # Redirect back to the hierarchy page with a success message
    messages.success(request, f'User {user.full_name} has been deleted successfully.')
    return redirect('user_hierarchy')
@login_required
def list_department_coordinators(request):
    coordinators = User.objects.filter(account_type='Department Coordinator')
    
    # Group by department
    grouped = defaultdict(list)
    for user in coordinators:
        department = user.department if user.department else 'No Department'
        grouped[department].append(user)

    return render(request, 'accounts/department_coordinators_list.html', {
        'grouped_coordinators': dict(grouped)
    })

@login_required
def list_extensionists(request):
    extensionists = User.objects.filter(account_type='Extensionist')

    grouped = defaultdict(list)
    for user in extensionists:
        department = user.department if user.department else 'No Department'
        grouped[department].append(user)

    return render(request, 'accounts/extensionists_list.html', {
        'grouped_extensionists': dict(grouped)
    })

@login_required
def super_admin_view(request):
    if request.user.account_type != AccountType.SUPER_ADMIN:
        return redirect('permission_denied')
    return render(request, 'accounts/super_admin_dashboard.html')

from django.contrib.auth.decorators import login_required
from django.utils import timezone
from datetime import timedelta
from django.shortcuts import redirect, render
from .models import Document, AccountType

# Common cutoff
cutoff_date = timezone.now() - timedelta(days=30)

from django.contrib.auth.decorators import login_required
from django.shortcuts import redirect, render
from django.utils.timezone import now, timedelta
from .models import Document, Department, AccountType

from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect
from django.utils.timezone import now
from datetime import timedelta
from .models import Document, Department, AccountType
 
@login_required
def campus_admin_view(request):
    if request.user.account_type != AccountType.CAMPUS_ADMIN:
        return redirect('permission_denied')

    cutoff_date = now() - timedelta(days=30)

    # ✅ Recent updates
    recent_updates = Document.objects.filter(
        status_updated_at__gte=cutoff_date
    ).order_by('-status_updated_at')[:20]

    # ✅ Status sections
    statuses = [
        ('pending', 'Pending'),
        ('recommended', 'Recommended'),
        ('approved', 'Approved'),

        ('ongoing', 'Ongoing'),
        ('completion_processing', 'Completion Reviewing'),
        ('completion_recommended', 'Completion Recommended'),
        ('completed', 'Completed'),
    ]

    status_data = []
    for code, label in statuses:
        docs = Document.objects.filter(
            status=code,
            status_updated_at__gte=cutoff_date
        ).order_by('-status_updated_at')[:10]
        status_data.append({'label': label, 'code': code, 'documents': docs})

    # ✅ Nicknames for document file labels
    INITIAL_NICKNAMES = {
        "Activity_Proposal": "Activity Proposal",
        "Work_and_Financial_Plan": "Work & Financial Plan",
        "Plan_of_Activities": "Plan of Activities",
        "doc4": "Supporting Document 1",
        "doc5": "Supporting Document 2",
        "doc6": "Supporting Document 3",
        "doc7": "Supporting Document 4",
        "doc8": "Supporting Document 5",
    }

    DAYFILE_NICKNAMES = {
        "doc1": "Attendance Sheet ",
        "doc2": "Photo Documentation",
        "doc3": "Program",
        "doc4": "Supporting Document 1",
        "doc5": "Supporting Document 2",
    }

    COMPLETION_NICKNAMES = {
        "completion_doc1": "Approved Letter Request",
        "completion_doc2": "Accomplished/Evaluation Form",
        "completion_doc3": "Supporting Document 1",
        "completion_doc4": "Supporting Document 2",
        "completion_doc5": "Supporting Document 3",
        "completion_doc6": "Supporting Document 4",
        "completion_doc7": "Supporting Document 5",
        "completion_doc8": "Supporting Document 6",
    }

    # ✅ Department revisions section
    departments = Department.objects.all()
    dept_data = []

    for dept in departments:
        revision_docs = []

        for doc in dept.document_set.all():
            revision_files = []

            # --- (1) Initial uploads on Document ---
            initial_fields = [
                "Activity_Proposal",
                "Work_and_Financial_Plan",
                "Plan_of_Activities"
            ] + [f"doc{i}" for i in range(4, 9)]

            for field in initial_fields:
                status_field = f"{field}_status"
                if hasattr(doc, status_field) and getattr(doc, status_field) == "revision":
                    nickname = INITIAL_NICKNAMES.get(field, field.replace("_", " ").title())
                    revision_files.append(nickname)

            # --- (2) Supporting files from DocumentDayFile ---
            for day in doc.days.all():
                for dayfile in day.day_files.all():
                    for field, nickname in DAYFILE_NICKNAMES.items():
                        status_field = f"{field}_status"
                        if hasattr(dayfile, status_field) and getattr(dayfile, status_field) == "revision":
                            revision_files.append(nickname)

            # --- (3) Completion uploads ---
            for completion in doc.files.all():
                for field, nickname in COMPLETION_NICKNAMES.items():
                    status_field = f"{field}_status"
                    if hasattr(completion, status_field) and getattr(completion, status_field) == "revision":
                        revision_files.append(nickname)

            # ✅ Only include docs with at least one revision
            if revision_files:
                revision_docs.append({
                    "id": doc.id,
                    "name": doc.name,
                    "revision_files": revision_files
                })

        dept_data.append({
            "id": dept.id,
            "name": dept.name,
            "documents_with_revisions": revision_docs,
        })
    moa_list = MOAResource.objects.all().order_by('-uploaded_at')
    # ✅ Render the template
    return render(request, 'accounts/campus_admin_dashboard.html', {
        'recent_updates': recent_updates,
        'status_data': status_data,
        'departments': dept_data,
        'moa_list': moa_list,
        'moa_form': MOAResourceForm(),
    })
@login_required
def upload_moa(request):
    if request.method == 'POST':
        form = MOAResourceForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
    return redirect('campus_admin_dashboard')
@login_required
def delete_moa(request, moa_id):
    moa = MOAResource.objects.get(id=moa_id)
    moa.delete()
    return redirect('campus_admin_dashboard')
@login_required
def edit_moa(request, moa_id):
    moa = MOAResource.objects.get(id=moa_id)
    if request.method == 'POST':
        form = MOAResourceForm(request.POST, request.FILES, instance=moa)
        if form.is_valid():
            form.save()
            return redirect('campus_admin_dashboard')

    return render(request, 'accounts/edit_moa.html', {'form': MOAResourceForm(instance=moa)})

from django.utils.timezone import now
from datetime import timedelta
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect
from .models import Document, Department, AccountType

# -------------------------
# Staff Extensionist View
# -------------------------
@login_required
def staff_extensionist_view(request):
    if request.user.account_type != AccountType.STAFF_EXTENSIONIST:
        return redirect('permission_denied')

    cutoff_date = now() - timedelta(days=30)

    # Recent updates
    recent_updates = Document.objects.filter(
        status_updated_at__gte=cutoff_date
    ).order_by('-status_updated_at')[:20]

    # Status sections
    statuses = [
        ('pending', 'Pending'),
        ('recommended', 'Recommended'),
        ('approved', 'Approved'),

        ('ongoing', 'Ongoing'),
        ('completion_processing', 'Completion Reviewing'),
        ('completion_recommended', 'Completion Recommended'),
        ('completed', 'Completed'),
    ]

    status_data = []
    for code, label in statuses:
        docs = Document.objects.filter(
            status=code,
            status_updated_at__gte=cutoff_date
        ).order_by('-status_updated_at')[:10]
        status_data.append({'label': label, 'code': code, 'documents': docs})

    # ✅ Nicknames for document file labels
    INITIAL_NICKNAMES = {
        "Activity_Proposal": "Activity Proposal",
        "Work_and_Financial_Plan": "Work & Financial Plan",
        "Plan_of_Activities": "Plan of Activities",
        "doc4": "Supporting Document 1",
        "doc5": "Supporting Document 2",
        "doc6": "Supporting Document 3",
        "doc7": "Supporting Document 4",
        "doc8": "Supporting Document 5",
    }

    DAYFILE_NICKNAMES = {
        "doc1": "Attendance Sheet ",
        "doc2": "Photo Documentation",
        "doc3": "Program",
        "doc4": "Supporting Document 1",
        "doc5": "Supporting Document 2",
    }

    COMPLETION_NICKNAMES = {
        "completion_doc1": "Approved Letter Request",
        "completion_doc2": "Accomplished/Evaluation Form",
        "completion_doc3": "Supporting Document 1",
        "completion_doc4": "Supporting Document 2",
        "completion_doc5": "Supporting Document 3",
        "completion_doc6": "Supporting Document 4",
        "completion_doc7": "Supporting Document 5",
        "completion_doc8": "Supporting Document 6",
    }

    # Department revisions
    departments = Department.objects.all()
    dept_data = []

    for dept in departments:
        revision_docs = []

        for doc in dept.document_set.all():
            revision_files = []

            # Initial uploads
            initial_fields = ["Activity_Proposal", "Work_and_Financial_Plan", "Plan_of_Activities"] + [f"doc{i}" for i in range(4, 9)]
            for field in initial_fields:
                status_field = f"{field}_status"
                if hasattr(doc, status_field) and getattr(doc, status_field) == "revision":
                    revision_files.append(INITIAL_NICKNAMES.get(field, field.replace("_", " ").title()))

            # Supporting documents
            for day in doc.days.all():
                for dayfile in day.day_files.all():
                    for field, nickname in DAYFILE_NICKNAMES.items():
                        status_field = f"{field}_status"
                        if hasattr(dayfile, status_field) and getattr(dayfile, status_field) == "revision":
                            revision_files.append(nickname)

            # Completion uploads
            for completion in doc.files.all():
                for field, nickname in COMPLETION_NICKNAMES.items():
                    status_field = f"{field}_status"
                    if hasattr(completion, status_field) and getattr(completion, status_field) == "revision":
                        revision_files.append(nickname)

            if revision_files:
                revision_docs.append({
                    "id": doc.id,
                    "name": doc.name,
                    "revision_files": revision_files
                })

        dept_data.append({
            "id": dept.id,
            "name": dept.name,
            "documents_with_revisions": revision_docs,
        })
        moa_list = MOAResource.objects.all().order_by('-uploaded_at')

    return render(request, 'accounts/staff_extensionist_dashboard.html', {
        'recent_updates': recent_updates,
        'status_data': status_data,
        'departments': dept_data,
        'moa_list': moa_list,
    })


# -------------------------
# Department Coordinator View
# -------------------------
@login_required
def department_coordinator_view(request):
    if request.user.account_type != AccountType.DEPARTMENT_COORDINATOR:
        return redirect('permission_denied')
    if not getattr(request.user, 'department', None):
        return redirect('permission_denied')

    department = request.user.department
    cutoff_date = now() - timedelta(days=30)

    # Recent updates
    recent_updates = Document.objects.filter(
        department=department,
        status_updated_at__gte=cutoff_date
    ).order_by('-status_updated_at')[:20]

    # Status sections
    statuses = [
        ('pending', 'Pending'),
        ('recommended', 'Recommended'),
        ('approved', 'Approved'),

        ('ongoing', 'Ongoing'),
        ('completion_processing', 'Completion Reviewing'),
        ('completion_recommended', 'Completion Recommended'),
        ('completed', 'Completed'),
    ]

    status_data = []
    for code, label in statuses:
        docs = Document.objects.filter(
            department=department,
            status=code,
            status_updated_at__gte=cutoff_date
        ).order_by('-status_updated_at')[:10]
        status_data.append({'label': label, 'code': code, 'documents': docs})

    # ✅ Nicknames for document file labels
    INITIAL_NICKNAMES = {
        "Activity_Proposal": "Activity Proposal",
        "Work_and_Financial_Plan": "Work & Financial Plan",
        "Plan_of_Activities": "Plan of Activities",
        "doc4": "Supporting Document 1",
        "doc5": "Supporting Document 2",
        "doc6": "Supporting Document 3",
        "doc7": "Supporting Document 4",
        "doc8": "Supporting Document 5",
    }

    DAYFILE_NICKNAMES = {
        "doc1": "Attendance Sheet ",
        "doc2": "Photo Documentation",
        "doc3": "Program",
        "doc4": "Supporting Document 1",
        "doc5": "Supporting Document 2",
    }

    COMPLETION_NICKNAMES = {
        "completion_doc1": "Approved Letter Request",
        "completion_doc2": "Accomplished/Evaluation Form",
        "completion_doc3": "Supporting Document 1",
        "completion_doc4": "Supporting Document 2",
        "completion_doc5": "Supporting Document 3",
        "completion_doc6": "Supporting Document 4",
        "completion_doc7": "Supporting Document 5",
        "completion_doc8": "Supporting Document 6",
    }

    # Documents with revisions
    dept_documents = Document.objects.filter(department=department).prefetch_related("files", "days__day_files")
    documents_with_revisions = []

    for doc in dept_documents:
        revision_files = []

        # Initial uploads
        initial_fields = ["Activity_Proposal", "Work_and_Financial_Plan", "Plan_of_Activities"] + [f"doc{i}" for i in range(4, 9)]
        for field in initial_fields:
            status_field = f"{field}_status"
            if hasattr(doc, status_field) and getattr(doc, status_field) == "revision":
                revision_files.append(INITIAL_NICKNAMES.get(field, field.replace("_", " ").title()))

        # Supporting files
        for day in doc.days.all():
            for dayfile in day.day_files.all():
                for field, nickname in DAYFILE_NICKNAMES.items():
                    status_field = f"{field}_status"
                    if hasattr(dayfile, status_field) and getattr(dayfile, status_field) == "revision":
                        revision_files.append(nickname)

        # Completion uploads
        for completion in doc.files.all():
            for field, nickname in COMPLETION_NICKNAMES.items():
                status_field = f"{field}_status"
                if hasattr(completion, status_field) and getattr(completion, status_field) == "revision":
                    revision_files.append(nickname)

        if revision_files:
            documents_with_revisions.append({
                "id": doc.id,
                "name": doc.name,
                "revision_files": revision_files
            })
        moa_list = MOAResource.objects.all().order_by('-uploaded_at')

    return render(request, 'accounts/department_coordinator_dashboard.html', {
        'recent_updates': recent_updates,
        'status_data': status_data,
        'departments': [{
            "id": department.id,
            "name": department.get_name_display(),
            "documents_with_revisions": documents_with_revisions
        }],
        'moa_list': moa_list,
    })


# -------------------------
# Extensionist View
# -------------------------
@login_required
def extensionist_view(request):
    if request.user.account_type != AccountType.EXTENSIONIST:
        return redirect('permission_denied')
    if not getattr(request.user, 'department', None):
        return redirect('permission_denied')

    department = request.user.department
    cutoff_date = now() - timedelta(days=30)

    # Recent updates
    recent_updates = Document.objects.filter(
        department=department,
        status_updated_at__gte=cutoff_date
    ).order_by('-status_updated_at')[:20]

    # Status sections
    statuses = [
        ('pending', 'Pending'),
        ('recommended', 'Recommended'),
        ('approved', 'Approved'),

        ('ongoing', 'Ongoing'),
        ('completion_processing', 'Completion Reviewing'),
        ('completion_recommended', 'Completion Recommended'),
        ('completed', 'Completed'),
    ]

    status_data = []
    for code, label in statuses:
        docs = Document.objects.filter(
            department=department,
            status=code,
            status_updated_at__gte=cutoff_date
        ).order_by('-status_updated_at')[:10]
        status_data.append({'label': label, 'code': code, 'documents': docs})

    # Nicknames
    INITIAL_NICKNAMES = {
        "Activity_Proposal": "Activity Proposal",
        "Work_and_Financial_Plan": "Work & Financial Plan",
        "Plan_of_Activities": "Plan of Activities",
        "doc4": "Invitation Letter",
        "doc5": "Pre-test/Post-test",
        "doc6": "Evaluation Form",
        "doc7": "Accomplishment Report",
        "doc8": "Supporting Document",
    }

    DAYFILE_NICKNAMES = {
        "doc1": "Supporting Document 1",
        "doc2": "Supporting Document 2",
        "doc3": "Supporting Document 3",
        "doc4": "Supporting Document 4",
        "doc5": "Supporting Document 5",
    }

    COMPLETION_NICKNAMES = {
        "completion_doc1": "Activity Report",
        "completion_doc2": "Training Summary",
        "completion_doc3": "Attendance Sheet",
        "completion_doc4": "Photos",
        "completion_doc5": "Certificate of Completion",
        "completion_doc6": "Evaluation Summary",
        "completion_doc7": "Financial Report",
        "completion_doc8": "MOA/Partner Document",
    }

    # Documents with revisions
    dept_documents = Document.objects.filter(department=department).prefetch_related("files", "days__day_files")
    documents_with_revisions = []

    for doc in dept_documents:
        revision_files = []

        # Initial uploads
        initial_fields = ["Activity_Proposal", "Work_and_Financial_Plan", "Plan_of_Activities"] + [f"doc{i}" for i in range(4, 9)]
        for field in initial_fields:
            status_field = f"{field}_status"
            if hasattr(doc, status_field) and getattr(doc, status_field) == "revision":
                revision_files.append(INITIAL_NICKNAMES.get(field, field.replace("_", " ").title()))

        # Supporting files
        for day in doc.days.all():
            for dayfile in day.day_files.all():
                for field, nickname in DAYFILE_NICKNAMES.items():
                    status_field = f"{field}_status"
                    if hasattr(dayfile, status_field) and getattr(dayfile, status_field) == "revision":
                        revision_files.append(nickname)

        # Completion uploads
        for completion in doc.files.all():
            for field, nickname in COMPLETION_NICKNAMES.items():
                status_field = f"{field}_status"
                if hasattr(completion, status_field) and getattr(completion, status_field) == "revision":
                    revision_files.append(nickname)

        if revision_files:
            documents_with_revisions.append({
                "id": doc.id,
                "name": doc.name,
                "revision_files": revision_files
            })

    return render(request, 'accounts/extensionist_dashboard.html', {
        'recent_updates': recent_updates,
        'status_data': status_data,
        'departments': [{
            "id": department.id,
            "name": department.get_name_display(),
            "documents_with_revisions": documents_with_revisions
        }]
    })


from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect
from django.contrib import messages

# Permission Denied View
def permission_denied(request):
    return render(request, 'accounts/permission_denied.html')

@login_required
def upload_document(request):

    # Allow only Department Coordinator and Staff Extensionist
    allowed_roles = [
        AccountType.DEPARTMENT_COORDINATOR,
        AccountType.STAFF_EXTENSIONIST,
    ]

    if request.user.account_type not in allowed_roles:
        return redirect('permission_denied')

    # Department is optional for Staff Extensionist but required for Coordinator
    if request.user.account_type == AccountType.DEPARTMENT_COORDINATOR:
        if not getattr(request.user, 'department', None):
            return redirect('permission_denied')

    print(f"[DEBUG] User: {request.user.username} ({request.user.account_type})")

    if request.method == 'POST':
        form = DocumentUploadForm(request.POST, request.FILES)
        print(f"[DEBUG] Form valid: {form.is_valid()}")
        print(f"[DEBUG] Form errors: {form.errors}")

        if form.is_valid():
            document = form.save(commit=False)
            document.uploaded_by = request.user
            document.uploaded_by_name = request.user.get_full_name() or request.user.username

            # Department for Coordinator; optional for Staff Extensionist
            document.department = getattr(request.user, 'department', None)

            # If no department is attached, mark document as public
            document.is_public = document.department is None

            document.save()

            # -------- EMAIL NOTIFICATION --------
            try:
                from django.core.mail import send_mail
                from django.conf import settings

                subject = f"📄 New Document Uploaded: {document.name}"
                message = f"""
Hi Staff Extensionist,

Uploaded by: {document.uploaded_by.username} ({document.uploaded_by.email})

Title: {document.name}
Department: {document.department.name if document.department else 'N/A'}

Please review it in the system.

Best,
The System ✨
"""

                recipients = [
                    user.email
                    for user in CustomUser.objects.filter(
                        account_type=AccountType.STAFF_EXTENSIONIST
                    )
                    if user.email
                ]

                if recipients:
                    send_mail(
                        subject,
                        message,
                        settings.DEFAULT_FROM_EMAIL,
                        recipients,
                        fail_silently=False,
                    )
            except Exception as e:
                print("[DEBUG] Email failed:", e)

            messages.success(request, "Document uploaded successfully.")
            return redirect('pending_documents')

    else:
        form = DocumentUploadForm()

    return render(request, 'accounts/upload_document.html', {'form': form})





# List pending documents

@login_required
def pending_documents(request):
    user = request.user

    if user.account_type in [AccountType.DEPARTMENT_COORDINATOR, AccountType.EXTENSIONIST]:
        documents = Document.objects.filter(
            status='pending'
        ).filter(
            Q(department=user.department) |
            Q(uploaded_by__account_type=AccountType.STAFF_EXTENSIONIST)
        ).distinct().order_by('-created_at')  # Use existing timestamp field
    else:
        documents = Document.objects.filter(status='pending').order_by('-created_at')

    return render(request, 'accounts/pending_documents.html', {'documents': documents})


@login_required
def ongoing_documents(request):
    user = request.user

    if user.account_type in [AccountType.DEPARTMENT_COORDINATOR, AccountType.EXTENSIONIST]:
        documents = Document.objects.filter(
            status='ongoing'
        ).filter(
            Q(department=user.department) |
            Q(uploaded_by__account_type=AccountType.STAFF_EXTENSIONIST)
        ).distinct().order_by('-created_at')
    else:
        documents = Document.objects.filter(status='ongoing').order_by('-created_at')

    return render(request, 'accounts/ongoing_documents.html', {'documents': documents})

# Staff Extensionist recommends a document
@login_required
def recommend_document(request, doc_id):
    doc = get_object_or_404(Document, id=doc_id)
    if request.user.account_type != 'Staff Extensionist':
        return HttpResponseForbidden()
    doc.status = 'recommended'  # Use string 'recommended' directly
    doc.save()
    return redirect('recommended_documents')

# Campus Admin approves
@login_required
def approve_document(request, doc_id):
    doc = get_object_or_404(Document, id=doc_id)
    if request.user.account_type != 'Campus Admin':
        return HttpResponseForbidden()
    doc.status = 'ongoing'  # Use string 'ongoing' directly
    doc.save()
    return redirect('recommended_documents')



# Campus Admin rejects
@login_required
def reject_document(request, doc_id):
    doc = get_object_or_404(Document, id=doc_id)
    user = request.user

    # Super restricted access if not Campus Admin or other permitted roles
    if user.account_type == AccountType.CAMPUS_ADMIN:
        pass  # Campus Admins can always reject
    elif user.account_type in [AccountType.DEPARTMENT_COORDINATOR, AccountType.EXTENSIONIST]:
        # Must be in same department OR uploaded by staff extensionist
        if not (
            doc.department == user.department or 
            doc.uploaded_by.account_type == AccountType.STAFF_EXTENSIONIST
        ):
            return HttpResponseForbidden()
    else:
        return HttpResponseForbidden()

    # Reject the document
    doc.status = 'rejected'
    doc.save()
    return redirect('recommended_documents')
@login_required
def need_revision_document(request, doc_id):
    doc = get_object_or_404(Document, id=doc_id)
    user = request.user

    # Access control (same as reject)
    if user.account_type == AccountType.CAMPUS_ADMIN:
        pass
    elif user.account_type in [AccountType.DEPARTMENT_COORDINATOR, AccountType.EXTENSIONIST]:
        if not (
            doc.department == user.department or 
            doc.uploaded_by.account_type == AccountType.STAFF_EXTENSIONIST
        ):
            return HttpResponseForbidden()
    else:
        return HttpResponseForbidden()

    # Set document status back to "pending"
    doc.status = 'pending'
    doc.save()
    return redirect('recommended_documents')

# List recommended documents
@login_required
def recommended_documents(request):
    documents = Document.objects.filter(status='recommended').order_by('-created_at')
    return render(request, 'accounts/recommended_documents.html', {'documents': documents})


# Rejected documents view
@login_required
def rejected_documents(request):
    documents = Document.objects.filter(status='rejected').order_by('-created_at')
    return render(request, 'accounts/rejected_documents.html', {'documents': documents})

from django.shortcuts import render, get_object_or_404, redirect
from .models import Document, DocumentFile
from .forms import CompletionUploadForm
from django.contrib.auth.decorators import login_required

@login_required
def completion_upload(request, document_id):
    document = get_object_or_404(Document, id=document_id)
    document_file, _ = DocumentFile.objects.get_or_create(document=document)

    if request.method == 'POST':
        form = CompletionUploadForm(request.POST, request.FILES)
        if form.is_valid():
            slot = form.cleaned_data['slot_choice']
            file = form.cleaned_data['file']

            # Dynamically assign file to the selected slot
            setattr(document_file, slot, file)
            document_file.save()

            # Update document status
            document.status = 'completion_processing'
            document.save()

            return redirect('completion_pending_documents')
    else:
        form = CompletionUploadForm()

    return render(request, 'accounts/upload_completion.html', {
        'form': form,
        'existing_completion': document_file,
    })

@login_required
def completion_pending_documents(request):
    user = request.user

    if user.account_type in [AccountType.DEPARTMENT_COORDINATOR, AccountType.EXTENSIONIST]:
        documents = Document.objects.filter(
            status='completion_processing'
        ).filter(
            Q(department=user.department) |
            Q(uploaded_by__account_type=AccountType.STAFF_EXTENSIONIST)
        ).distinct().order_by('-created_at')
    else:
        documents = Document.objects.filter(status='completion_processing').order_by('-created_at')

    return render(request, 'accounts/completion_pending_documents.html', {'documents': documents})

@login_required
def recommend_completion(request, doc_id):
    if request.user.account_type != 'Staff Extensionist':
        return HttpResponseForbidden()
    doc = get_object_or_404(Document, id=doc_id)
    doc.status = 'completion_recommended'
    doc.recommended_by = request.user
    doc.save()
    return redirect('completion_pending_documents')

@login_required
def reject_completion(request, doc_id):
    if request.user.account_type not in ['Staff Extensionist', 'Campus Admin']:
        return HttpResponseForbidden("You do not have permission to reject this document.")
    
    doc = get_object_or_404(Document, id=doc_id)
    doc.status = 'ongoing'  # Revert the status to ongoing
    doc.save()
    return redirect('completion_pending_documents')

@login_required
def approve_completion(request, doc_id):
    if request.user.account_type != 'Campus Admin':
        return HttpResponseForbidden()
    doc = get_object_or_404(Document, id=doc_id)
    if doc.status == 'completion_recommended':
        doc.status = 'completed'
        doc.approved_by = request.user
        doc.save()
    return redirect('completion_recommended_documents')
@login_required
def completion_recommended_documents(request):
    documents = Document.objects.filter(status='completion_recommended').order_by('-created_at')
    return render(request, 'accounts/completion_recommended_documents.html', {'documents': documents})



# Completed documents view
@login_required
def completed_documents(request):
    documents = Document.objects.filter(status='completed').order_by('-created_at')
    return render(request, 'accounts/completed_documents.html', {'documents': documents})

@login_required
def document_list(request):
    user = request.user
    selected_status = request.GET.get('status')

    if user.account_type in [AccountType.DEPARTMENT_COORDINATOR, AccountType.EXTENSIONIST]:
        documents = Document.objects.filter(
            Q(department=user.department) |
            Q(uploaded_by__account_type=AccountType.STAFF_EXTENSIONIST)
        ).distinct()
    else:
        documents = Document.objects.all()

    # ✅ Exclude documents where status is None or empty
    documents = documents.exclude(status__isnull=True).exclude(status='')

    if selected_status:
        documents = documents.filter(status=selected_status)

    documents = documents.order_by('-created_at')  # Sort newest first

    all_statuses = [
        ('pending', 'Pending'),
        ('ongoing', 'Ongoing'),
        ('completion_processing', 'Completion Reviewing'),
        ('completed', 'Completed'),
    ]

    if user.account_type in [AccountType.SUPER_ADMIN, AccountType.CAMPUS_ADMIN]:
        all_statuses += [
            ('completion_recommended', 'Completion Recommended'),
            ('recommended', 'Recommended'),
        ]

    context = {
        'documents': documents,
        'selected_status': selected_status,
        'all_statuses': all_statuses,
    }
    return render(request, 'accounts/document_list.html', context)

from django.shortcuts import get_object_or_404, redirect, render
from django.contrib.auth.decorators import login_required
from django.http import HttpResponseForbidden
from django.contrib import messages
import os

# -------------------------------
# HELPER FUNCTIONS
# -------------------------------
def clear_feedback(document):
    DocumentRevisionFeedback.objects.filter(document=document).delete()
    CompletionRevisionFeedback.objects.filter(document=document).delete()
    DayRevisionFeedback.objects.filter(document=document).delete()

def get_post_id(request, name):
    val = request.POST.get(name)
    try:
        return int(val)
    except (TypeError, ValueError):
        return None

def handle_file_upload(instance, slot_name, uploaded_file):
    setattr(instance, slot_name, uploaded_file)
    status_field = f"{slot_name}_status"
    if hasattr(instance, status_field):
        setattr(instance, status_field, "normal")
    instance.save()
    
from django.utils import timezone

def handle_file_upload(instance, slot_choice, uploaded_file, user=None):
    """
    Handle file upload and track uploader info
    """
    if hasattr(instance, slot_choice):
        setattr(instance, slot_choice, uploaded_file)

        # Store uploader info if available
        if user:
            uploader_field = f"{slot_choice}_uploaded_by"
            timestamp_field = f"{slot_choice}_uploaded_at"
            status_field = f"{slot_choice}_status"

            if hasattr(instance, uploader_field):
                setattr(instance, uploader_field, user)

            if hasattr(instance, timestamp_field):
                setattr(instance, timestamp_field, timezone.now())

            # NEW: auto set status so NOT NULL constraint is satisfied
            if hasattr(instance, status_field):
                setattr(instance, status_field, "uploaded")   # or pending, draft, etc.

        instance.save()

# -------------------------------
# VIEW
# -------------------------------
@login_required
def view_document(request, document_id):
    document = get_object_or_404(Document, id=document_id)
    user = request.user

    # -------------------------------
    # ACCESS CONTROL: VIEW DOCUMENT
    # -------------------------------
    super_roles = ["Campus Admin", "Staff Extensionist"]
    dept_roles = ["Department Coordinator", "Extensionist"]

    if user.account_type in super_roles:
        pass
    elif user.account_type in dept_roles:
        if document.department != getattr(user, "department", None):
            return redirect("permission_denied")
    else:
        return redirect("permission_denied")


    # Permissions
    can_upload_completion = user.account_type in ["Department Coordinator", "Extensionist", "Staff Extensionist"] and document.status == "ongoing"
    can_comment = user.account_type in ["Campus Admin", "Staff Extensionist"]

    # Forms & Data
    existing_completion = DocumentFile.objects.filter(document=document).first()
    form = CompletionUploadForm(instance=existing_completion)
    edit_initial_form = DocumentUploadForm(instance=document)
    comment_form = DocumentCommentForm()
    day_form = DocumentDayForm()
    report_form = DayTrainingReportForm()
    comments = DocumentComment.objects.filter(document=document).order_by("-created_at")
    revision_feedbacks = DocumentRevisionFeedback.objects.filter(document=document).order_by("-created_at")
    day_revision_feedbacks = DayRevisionFeedback.objects.filter(document=document)
    completion_feedbacks = CompletionRevisionFeedback.objects.filter(document=document)

    if request.method == "POST":

        # -----------------------
        # DELETE DOCUMENT
        # -----------------------
        if "delete_document" in request.POST:
            if user == document.uploaded_by or user.account_type in ["Campus Admin", "Staff Extensionist", "Department Coordinator"]:
                document.is_archived = True
                document.status = None
                document.save()
                return redirect("archived_documents")
            return HttpResponseForbidden("You do not have permission to delete this document.")

        # -----------------------
        # COMPLETION UPLOAD
        # -----------------------
        elif "upload_completion" in request.POST:
            slot_choice = request.POST.get("completion_slot_choice")
            uploaded_file = request.FILES.get("selected_file")

            # Only upload file if a slot + file are provided
            if slot_choice and uploaded_file:
                if not existing_completion:
                    existing_completion = DocumentFile(document=document)

                handle_file_upload(existing_completion, slot_choice, uploaded_file)

            # Status will still update even with no slot / no file
            document.status = "completion_processing"
            document.save()

            clear_feedback(document)
            messages.success(request, "Completion submitted. Now under processing.")
            return redirect("view_document", document_id=document.id)



        elif "save_completion_changes" in request.POST and can_upload_completion:
            slot_choice = request.POST.get("completion_slot_choice")
            uploaded_file = request.FILES.get("selected_file")
            valid_slots = [f"completion_doc{i}" for i in range(1, 9)]

            if slot_choice in valid_slots and uploaded_file:
                if not existing_completion:
                    existing_completion = DocumentFile(document=document)
                handle_file_upload(existing_completion, slot_choice, uploaded_file, user)  # ← Added user
                clear_feedback(document)
                messages.success(request, f"{slot_choice.replace('_', ' ').title()} updated successfully.")
            else:
                messages.error(request, "Invalid slot or file.")
            return redirect("view_document", document_id=document.id)
        
        elif "mark_revision_completion" in request.POST and can_comment:
            slot_choice = request.POST.get("revision_doc")
            valid_slots = [f"completion_doc{i}" for i in range(1, 9)]
            if slot_choice in valid_slots and existing_completion and hasattr(existing_completion, slot_choice):
                setattr(existing_completion, f"{slot_choice}_status", "revision")
                existing_completion.save()
                messages.warning(request, f"{slot_choice.replace('_', ' ').title()} marked for revision.")
            else:
                messages.error(request, "Invalid slot or completion document not found.")
            return redirect("view_document", document_id=document.id)

        elif "submit_completion_revision_comment" in request.POST and can_comment:
            slot_name = request.POST.get("completion_slot")
            comment_text = request.POST.get("revision_comment")
            image = request.FILES.get("revision_image")
            if slot_name and comment_text and existing_completion:
                CompletionRevisionFeedback.objects.create(
                    document=document,
                    completion_file=existing_completion,
                    slot_name=slot_name,
                    user=user,
                    comment=comment_text,
                    image=image
                )
                messages.success(request, "Completion document revision comment added successfully.")
            else:
                messages.error(request, "Missing comment text or invalid slot.")
            return redirect("view_document", document_id=document.id)

        # -----------------------
        # INITIAL UPLOAD
        # -----------------------
        elif "submit_initial_upload" in request.POST:
            slot_choice = request.POST.get("initial_slot_choice")
            uploaded_file = request.FILES.get("initial_selected_file")
            valid_slots = ["Activity_Proposal", "Work_and_Financial_Plan", "Plan_of_Activities"] + [f"doc{i}" for i in range(4, 9)]
            if slot_choice in valid_slots and uploaded_file:
                handle_file_upload(document, slot_choice, uploaded_file, user)  # ← Added user
                clear_feedback(document)
                messages.success(request, f"{slot_choice.replace('_', ' ').title()} uploaded successfully.")
            else:
                messages.error(request, "Invalid slot or file.")
            return redirect("view_document", document_id=document.id)

        elif "mark_initial_revision" in request.POST and can_comment:
            slot_choice = request.POST.get("revision_doc")
            comment_text = request.POST.get("revision_comment")
            image = request.FILES.get("revision_image")

            valid_slots = ["Activity_Proposal", "Work_and_Financial_Plan", "Plan_of_Activities"] + [f"doc{i}" for i in range(4, 9)]
            if slot_choice not in valid_slots:
                messages.error(request, "Invalid slot selected.")
                return redirect("view_document", document_id=document.id)

            # Mark the document slot for revision
            setattr(document, f"{slot_choice}_status", "revision")
            document.save()

            # Add comment if provided
            if comment_text:
                DocumentRevisionFeedback.objects.create(
                    document=document,
                    slot_name=slot_choice,
                    user=user,
                    comment=comment_text,
                    image=image
                )
                messages.warning(request, f"{slot_choice.replace('_', ' ').title()} marked for revision with feedback.")
            else:
                messages.warning(request, f"{slot_choice.replace('_', ' ').title()} marked for revision (no comment added).")

            return redirect("view_document", document_id=document.id)


        elif "submit_revision_comment" in request.POST:
            slot_choice = request.POST.get("slot_name")
            comment_text = request.POST.get("revision_comment")
            image = request.FILES.get("revision_image")
            if slot_choice and comment_text:
                DocumentRevisionFeedback.objects.create(
                    document=document,
                    slot_name=slot_choice,
                    user=user,
                    comment=comment_text,
                    image=image
                )
                messages.success(request, "Revision comment added successfully.")
            else:
                messages.error(request, "Missing comment text or invalid slot.")
            return redirect("view_document", document_id=document.id)

        # -----------------------
        # COMMENTS
        # -----------------------
        elif "submit_comment" in request.POST and can_comment:
            comment_form = DocumentCommentForm(request.POST)
            if comment_form.is_valid():
                new_comment = comment_form.save(commit=False)
                new_comment.document = document
                new_comment.user = user
                new_comment.save()
                messages.success(request, "Comment added.")
            return redirect("view_document", document_id=document.id)

        # -----------------------
        # DAYS
        # -----------------------
        elif "add_day" in request.POST:
            day_form = DocumentDayForm(request.POST)
            if day_form.is_valid():
                new_day = day_form.save(commit=False)
                new_day.document = document
                new_day.save()
                messages.success(request, "Day added successfully.")
            return redirect("view_document", document_id=document.id)

        elif "edit_day" in request.POST:
            day_id = get_post_id(request, "day_id")
            day = get_object_or_404(DocumentDay, id=day_id, document=document)
            form = DocumentDayForm(request.POST, instance=day)
            if form.is_valid():
                form.save()
                messages.success(request, "Day updated successfully.")
            return redirect("view_document", document_id=document.id)

        elif "delete_day" in request.POST:
            day_id = get_post_id(request, "day_id")
            day = get_object_or_404(DocumentDay, id=day_id, document=document)
            day.delete()
            messages.success(request, "Day deleted successfully.")
            return redirect("view_document", document_id=document.id)

        # -----------------------
        # DAY REPORTS
        # -----------------------
        elif "upload_day_report" in request.POST:
            day_id = get_post_id(request, "day_id")
            day = get_object_or_404(DocumentDay, id=day_id, document=document)
            uploaded_file = request.FILES.get("report_file")
            title = request.POST.get("title") or f"Report for {day.name}"

            if day.training_reports.exists():
                messages.error(request, "This day already has a training report.")
                return redirect("view_document", document_id=document.id)

            if uploaded_file and os.path.splitext(uploaded_file.name)[1].lower() not in [".doc", ".docx"]:
                messages.error(request, "Only .doc or .docx files allowed.")
                return redirect("view_document", document_id=document.id)

            DayTrainingReport.objects.create(
                day=day,
                title=title,
                file=uploaded_file,
                uploaded_by=user
            )
            messages.success(request, "Training report uploaded successfully.")
            return redirect("view_document", document_id=document.id)


        elif "edit_report" in request.POST:
            report_id = get_post_id(request, "report_id")
            report = get_object_or_404(DayTrainingReport, id=report_id)
            form = DayTrainingReportForm(request.POST, request.FILES, instance=report)
            if form.is_valid():
                form.save()
                messages.success(request, "Report updated successfully.")
            return redirect("view_document", document_id=document.id)

        elif "delete_report" in request.POST:
            report_id = get_post_id(request, "report_id")
            report = get_object_or_404(DayTrainingReport, id=report_id)
            report.delete()
            messages.success(request, "Report deleted successfully.")
            return redirect("view_document", document_id=document.id)

        # -----------------------
        # DAY FILES
        # -----------------------
        elif "upload_day_files" in request.POST:
            day_id = get_post_id(request, "day_id")
            slot = request.POST.get("file_slot")
            uploaded_file = request.FILES.get("selected_file")
            valid_slots = ["doc1", "doc2", "doc3", "doc4", "doc5"]

            if not day_id or slot not in valid_slots:
                messages.error(request, "Invalid day or slot.")
                return redirect("view_document", document_id=document.id)

            day = get_object_or_404(DocumentDay, id=day_id, document=document)
            day_files, _ = DocumentDayFile.objects.get_or_create(day=day)

            # ← Pass user here
            handle_file_upload(day_files, slot, uploaded_file, user)

            clear_feedback(document)
            messages.success(request, f"{slot.capitalize()} uploaded successfully for {day.title}.")
            return redirect("view_document", document_id=document.id)


        elif "delete_day_files" in request.POST:
            day_id = get_post_id(request, "day_id")
            day = get_object_or_404(DocumentDay, id=day_id, document=document)
            if day.day_files.exists():
                day.day_files.first().delete()
                messages.success(request, f"Supporting files for {day.title} deleted successfully.")
            else:
                messages.info(request, "No files found to delete.")
            return redirect("view_document", document_id=document.id)

        elif "make_day_revision" in request.POST and can_comment:
            day_id = get_post_id(request, "day_id")
            slot_choice = request.POST.get("revision_doc")
            valid_slots = [f"day_doc{i}" for i in range(1, 6)]

            if slot_choice not in valid_slots or not day_id:
                messages.error(request, "Invalid slot or day.")
                return redirect("view_document", document_id=document.id)

            day = get_object_or_404(DocumentDay, id=day_id, document=document)
            day_files = day.day_files.first()
            if day_files and hasattr(day_files, f"{slot_choice.replace('day_', '')}_status"):
                setattr(day_files, f"{slot_choice.replace('day_', '')}_status", "revision")
                day_files.save()
                messages.warning(request, f"{slot_choice.replace('day_', '').title()} marked for revision.")
            else:
                messages.error(request, "Invalid file slot.")
            return redirect("view_document", document_id=document.id)

        elif "submit_day_revision_comment" in request.POST:
            day_id = get_post_id(request, "day_id")
            slot_name = request.POST.get("slot_name")
            comment_text = request.POST.get("revision_comment")
            image = request.FILES.get("revision_image")

            if not day_id:
                messages.error(request, "Invalid day selected.")
                return redirect("view_document", document_id=document.id)

            day = get_object_or_404(DocumentDay, id=day_id, document=document)
            if slot_name and comment_text:
                DayRevisionFeedback.objects.create(
                    document=document,
                    day=day,
                    slot_name=slot_name,
                    user=user,
                    comment=comment_text,
                    image=image
                )
                messages.success(request, "Day revision comment added successfully.")
            else:
                messages.error(request, "Missing comment text or invalid slot.")
            return redirect("view_document", document_id=document.id)

    # --- Context & File Lists ---
    file_list = [
        ('Activity_Proposal', 'Activity Proposal', document.Activity_Proposal, document.Activity_Proposal_status,
        document.Activity_Proposal_uploaded_by, document.Activity_Proposal_uploaded_at),
        ('Work_and_Financial_Plan', 'Work and Financial Plan', document.Work_and_Financial_Plan, document.Work_and_Financial_Plan_status,
        document.Work_and_Financial_Plan_uploaded_by, document.Work_and_Financial_Plan_uploaded_at),
        ('Plan_of_Activities', 'Plan of Activities', document.Plan_of_Activities, document.Plan_of_Activities_status,
        document.Plan_of_Activities_uploaded_by, document.Plan_of_Activities_uploaded_at),
        ('doc4', 'Extra Document 1', document.doc4, document.doc4_status,
        document.doc4_uploaded_by, document.doc4_uploaded_at),
        ('doc5', 'Extra Document 2', document.doc5, document.doc5_status,
        document.doc5_uploaded_by, document.doc5_uploaded_at),
        ('doc6', 'Extra Document 3', document.doc6, document.doc6_status,
        document.doc6_uploaded_by, document.doc6_uploaded_at),
        ('doc7', 'Extra Document 4', document.doc7, document.doc7_status,
        document.doc7_uploaded_by, document.doc7_uploaded_at),
        ('doc8', 'Extra Document 5', document.doc8, document.doc8_status,
        document.doc8_uploaded_by, document.doc8_uploaded_at),
    ]

    if existing_completion:
        completion_file_list = [
            ('completion_doc1', 'Approve Letter Request', existing_completion.completion_doc1, existing_completion.completion_doc1_status,
            existing_completion.completion_doc1_uploaded_by, existing_completion.completion_doc1_uploaded_at),
            ('completion_doc2', 'Accomplished/Evaluation Form', existing_completion.completion_doc2, existing_completion.completion_doc2_status,
            existing_completion.completion_doc2_uploaded_by, existing_completion.completion_doc2_uploaded_at),
            ('completion_doc3', 'Extra Document 1', existing_completion.completion_doc3, existing_completion.completion_doc3_status,
            existing_completion.completion_doc3_uploaded_by, existing_completion.completion_doc3_uploaded_at),
            ('completion_doc4', 'Extra Document 2', existing_completion.completion_doc4, existing_completion.completion_doc4_status,
            existing_completion.completion_doc4_uploaded_by, existing_completion.completion_doc4_uploaded_at),
            ('completion_doc5', 'Extra Document 3', existing_completion.completion_doc5, existing_completion.completion_doc5_status,
            existing_completion.completion_doc5_uploaded_by, existing_completion.completion_doc5_uploaded_at),
            ('completion_doc6', 'Extra Document 4', existing_completion.completion_doc6, existing_completion.completion_doc6_status,
            existing_completion.completion_doc6_uploaded_by, existing_completion.completion_doc6_uploaded_at),
            ('completion_doc7', 'Extra Document 5', existing_completion.completion_doc7, existing_completion.completion_doc7_status,
            existing_completion.completion_doc7_uploaded_by, existing_completion.completion_doc7_uploaded_at),
            ('completion_doc8', 'Extra Document 6', existing_completion.completion_doc8, existing_completion.completion_doc8_status,
            existing_completion.completion_doc8_uploaded_by, existing_completion.completion_doc8_uploaded_at),
        ]
    else:
        completion_file_list = [
            ('completion_doc1', 'Approve Letter Request', None, None, None, None),
            ('completion_doc2', 'Accomplished/Evaluation Form', None, None, None, None),
            ('completion_doc3', 'Extra Document 1', None, None, None, None),
            ('completion_doc4', 'Extra Document 2', None, None, None, None),
            ('completion_doc5', 'Extra Document 3', None, None, None, None),
            ('completion_doc6', 'Extra Document 4', None, None, None, None),
            ('completion_doc7', 'Extra Document 5', None, None, None, None),
            ('completion_doc8', 'Extra Document 6', None, None, None, None),
        ]


    days = document.days.prefetch_related("training_reports")
    context = {
        "document": document,
        "form": form,
        "edit_initial_form": edit_initial_form,
        "existing_completion": existing_completion,
        "can_upload_completion": can_upload_completion,
        "can_comment": can_comment,
        "comment_form": comment_form,
        "comments": comments,
        "day_form": day_form,
        "report_form": report_form,
        "days": days,
        "day_file_form": DocumentDayFileForm(),
        "revision_feedback_form": RevisionFeedbackForm(),
        "file_list": file_list,
        "revision_feedbacks": revision_feedbacks,
        "day_revision_feedbacks": day_revision_feedbacks,
        'completion_file_list': completion_file_list,
        'completion_feedbacks': completion_feedbacks,
    }

    return render(request, "accounts/view_document.html", context)


from django.shortcuts import get_object_or_404, redirect
from django.contrib import messages
from django.http import HttpResponseForbidden
from .models import Document

@login_required
def archived_documents(request):
    user = request.user

    if 'recover_document' in request.POST:
        document_id = request.POST.get('document_id')
        document = get_object_or_404(Document, id=document_id)

        allowed_roles = ['Campus Admin', 'Staff Extensionist', 'Department Coordinator']
        if user.account_type not in allowed_roles:
            return HttpResponseForbidden("You do not have permission to recover this document.")

        document.is_archived = False
        document.status = 'pending'  # Set status back to 'pending' on recovery
        document.save()

        messages.success(request, "Document recovered successfully.")
        return redirect('archived_documents')

    elif 'delete_document' in request.POST:
        document_id = request.POST.get('document_id')
        document = get_object_or_404(Document, id=document_id)

        allowed_roles = ['Campus Admin', 'Staff Extensionist', 'Department Coordinator']
        if user.account_type not in allowed_roles:
            return HttpResponseForbidden("You do not have permission to delete this document.")

        document.delete()
        messages.success(request, "Document permanently deleted.")
        return redirect('archived_documents')

    # Fetch archived documents
    if user.account_type in [AccountType.DEPARTMENT_COORDINATOR, AccountType.EXTENSIONIST]:
        documents = Document.objects.filter(
            is_archived=True
        ).filter(
            Q(department=user.department) |
            Q(uploaded_by__account_type=AccountType.STAFF_EXTENSIONIST)
        ).distinct().order_by('-created_at')
    else:
        documents = Document.objects.filter(is_archived=True).order_by('-created_at')

    return render(request, 'accounts/archive_document.html', {'documents': documents})



from django.contrib.auth.decorators import login_required
from django.http import JsonResponse
from django.shortcuts import get_object_or_404
from django.utils.timezone import localtime
from .models import ChatMessage, Document
from django.utils.timezone import now
from django.utils.timezone import now

@login_required
def document_chat_list(request):
    user = request.user
    account_type = user.account_type
    department = user.department

    if account_type in [AccountType.SUPER_ADMIN, AccountType.CAMPUS_ADMIN, AccountType.STAFF_EXTENSIONIST]:
        documents = Document.objects.all()
    elif account_type in [AccountType.DEPARTMENT_COORDINATOR, AccountType.EXTENSIONIST]:
        documents = Document.objects.filter(department=department)
    else:
        documents = Document.objects.none()

    doc_data = []
    for doc in documents:
        latest_msg = ChatMessage.objects.filter(document=doc).order_by('-timestamp').first()
        last_updated = latest_msg.timestamp if latest_msg else doc.created_at or now()

        unread_count = ChatMessage.objects.filter(document=doc).exclude(read_by=user).count()

        doc_data.append({
            "id": doc.id,
            "name": doc.name,
            "unread": unread_count,
            "last_updated": last_updated.isoformat()
        })

    # ✅ Sort by recent message time, regardless of sender
    doc_data.sort(key=lambda x: x["last_updated"], reverse=True)

    return JsonResponse({"documents": doc_data})


@login_required
def post_chat_message(request, document_id):
    if request.method == 'POST':
        document = get_object_or_404(Document, id=document_id)
        message = request.POST.get('message', '').strip()

        if message:
            ChatMessage.objects.create(
                document=document,
                sender=request.user,
                sender_name_snapshot=request.user.get_full_name(),  # 👈 added
                message=message
            )
            return JsonResponse({'status': 'ok'})


        return JsonResponse({'status': 'empty message'}, status=400)

    return JsonResponse({'status': 'invalid method'}, status=405)

from django.utils.timezone import localtime
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse
from django.shortcuts import get_object_or_404
from .models import ChatMessage, Document

@login_required
def fetch_chat_messages(request, document_id):
    document = get_object_or_404(Document, id=document_id)
    messages = ChatMessage.objects.filter(document=document) \
                                  .select_related('sender') \
                                  .order_by('timestamp')

    data = []
    for msg in messages:
        localized_time = localtime(msg.timestamp)
        is_read = request.user in msg.read_by.all()
        data.append({
            'sender': msg.sender_name_snapshot or (msg.sender.get_full_name() if msg.sender else "Unknown User"),
            'sender_id': msg.sender.id if msg.sender else None,
            'message': msg.message,
            'timestamp': localtime(msg.timestamp).strftime('%b %d, %I:%M %p'),
            'read': request.user in msg.read_by.all(),
            'unread': request.user not in msg.read_by.all(),
        })

    return JsonResponse({
        'messages': data,
        'document_name': document.name  # ✅ Add this line
    })


from django.views.decorators.csrf import csrf_exempt

@csrf_exempt
@login_required
def mark_read_messages(request, document_id):
    if request.method == 'POST':
        document = get_object_or_404(Document, id=document_id)
        unread_messages = ChatMessage.objects.filter(document=document).exclude(read_by=request.user)
        for msg in unread_messages:
            msg.read_by.add(request.user)
        return JsonResponse({'status': 'ok'})

    return JsonResponse({'status': 'invalid method'}, status=405)



@login_required

def delete_coordinator(request, user_id):
    user = get_object_or_404(User, id=user_id)
    user.delete()
    messages.success(request, f'Coordinator {user.username} has been deleted.')
    return redirect('department_coordinators_list')


def create_annual_report(request):
    if request.method == 'POST':
        form = AnnualReportForm(request.POST)
        if form.is_valid():
            report = form.save()
            return redirect('view_report', report_id=report.id)
    else:
        form = AnnualReportForm()
    return render(request, 'accounts/create_annual_report.html', {'form': form})

@login_required
def view_report(request, report_id):
    report = get_object_or_404(AnnualReport, id=report_id)
    user = request.user

    # -------------------------------
    # ACCESS CONTROL
    # -------------------------------
    allowed_roles = ["Campus Admin", "Staff Extensionist"]

    if user.account_type not in allowed_roles:
        return redirect('permission_denied')

    # -------------------------------
    # ORIGINAL LOGIC
    # -------------------------------
    activities = report.activities.order_by('order')

    processed_activities = []
    activities_list = list(activities)

    i = 0
    while i < len(activities_list):
        current_act = activities_list[i]
        current_ext = current_act.extensionist

        span = 1
        for j in range(i + 1, len(activities_list)):
            if activities_list[j].extensionist == current_ext:
                span += 1
            else:
                break

        for k in range(span):
            processed_activities.append({
                'activity': activities_list[i + k],
                'show_extensionist': (k == 0),
                'rowspan': (span if k == 0 else 0)
            })

        i += span

    context = {
        'report': report,
        'processed_activities': processed_activities
    }
    return render(request, 'accounts/view_report.html', context)


@csrf_exempt
def reorder_activities(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        for index, activity_id in enumerate(data.get('activity_ids', [])):
            ExtensionActivity.objects.filter(id=activity_id).update(order=index)
        return JsonResponse({'status': 'ok'})
    return JsonResponse({'error': 'Invalid request'}, status=400)
@login_required
def add_activity(request, report_id):
    report = get_object_or_404(AnnualReport, id=report_id)
    if request.method == 'POST':
        form = ExtensionActivityForm(request.POST)
        if form.is_valid():
            activity = form.save(commit=False)
            activity.report = report
            activity.save()
            return redirect('view_report', report_id=report.id)
    else:
        form = ExtensionActivityForm()
    return render(request, 'accounts/add_activity.html', {'form': form, 'report': report})


@login_required
def generate_pdf(request, report_id):
    report = get_object_or_404(AnnualReport, id=report_id)
    activities = ExtensionActivity.objects.filter(report=report).order_by('order')
    linkages = report.linkages.all()

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    styles = getSampleStyleSheet()
    elements = []

    # Custom styles
    wrap_style = ParagraphStyle(name='Wrapped', fontSize=10, leading=12, alignment=TA_LEFT)
    white_heading = ParagraphStyle(name='WhiteHeading', parent=styles['Heading4'], textColor=colors.white)

    # Title
    elements.append(Paragraph(f"<b>ANNUAL REPORT {report.year}</b>", styles['Title']))
    elements.append(Spacer(1, 12))

    # Section I: Extension Activities
    elements.append(Paragraph("I. Extension", styles['Heading2']))
    elements.append(Paragraph("Table 27. List of extension activities conducted", styles['Normal']))
    elements.append(Spacer(1, 6))

    data = [[
        Paragraph("EXTENSION ACTIVITY", white_heading),
        Paragraph("EXTENSIONIST", white_heading),
        Paragraph("NO. OF CLIENTELE / BENEFICIARIES", white_heading),
        Paragraph("PARTNER AGENCY", white_heading)
    ]]

    span_commands = []

    activities_list = list(activities)
    i = 0
    row_index = 1  # start after header row

    while i < len(activities_list):
        current_ext = activities_list[i].extensionist
        start_row = row_index

        # Count how many consecutive activities share the same extensionist
        span = 1
        for j in range(i + 1, len(activities_list)):
            if activities_list[j].extensionist == current_ext:
                span += 1
            else:
                break

        for k in range(span):
            act = activities_list[i + k]
            data.append([
                Paragraph(act.activity.replace("\n", "<br/>"), wrap_style),
                Paragraph(current_ext, wrap_style) if k == 0 else '',
                Paragraph(act.no_of_beneficiaries, wrap_style),
                Paragraph(act.partner_agency, wrap_style)
            ])
            row_index += 1

        if span > 1:
            span_commands.append(('SPAN', (1, start_row), (1, row_index - 1)))

        i += span

    table = Table(data, colWidths=[140, 100, 130, 130])
    style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#8B0000")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('TOPPADDING', (0, 1), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ])
    for cmd in span_commands:
        style.add(*cmd)
    table.setStyle(style)

    elements.append(table)
    elements.append(Spacer(1, 20))

    # Google Drive Link
    if report.google_drive_link:
        elements.append(Paragraph(f"Google Drive Link: <i>{report.google_drive_link}</i>", styles['Normal']))
        elements.append(Spacer(1, 12))

    # Section II: Linkages
    elements.append(Paragraph("II. Linkages", styles['Heading2']))
    elements.append(Paragraph("Table 28. List of partner agencies and nature of linkages", styles['Normal']))
    elements.append(Spacer(1, 6))

    link_data = [[
        Paragraph("AGENCY", white_heading),
        Paragraph("NATURE OF LINKAGE", white_heading)
    ]]
    for link in linkages:
        link_data.append([
            Paragraph(link.agency, wrap_style),
            Paragraph(link.nature, wrap_style)
        ])

    link_table = Table(link_data, colWidths=[200, 300])
    link_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#8B0000")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('TOPPADDING', (0, 1), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    elements.append(link_table)

    # Generate PDF
    doc.build(elements)
    buffer.seek(0)
    return FileResponse(buffer, as_attachment=False, filename=f"Annual_Report_{report.year}.pdf")


@login_required
def edit_activity(request, activity_id):
    activity = get_object_or_404(ExtensionActivity, id=activity_id)
    if request.method == 'POST':
        form = ExtensionActivityForm(request.POST, instance=activity)
        if form.is_valid():
            form.save()
            return redirect('view_report', report_id=activity.report.id)
    else:
        form = ExtensionActivityForm(instance=activity)
    
    # Pass the activity context to the template as well
    return render(request, 'accounts/add_activity.html', {'form': form, 'activity': activity, 'report': activity.report})
@login_required
def edit_linkage(request, linkage_id):
    linkage = get_object_or_404(Linkage, id=linkage_id)
    if request.method == 'POST':
        form = LinkageForm(request.POST, instance=linkage)
        if form.is_valid():
            form.save()
            return redirect('view_report', report_id=linkage.report.id)
    else:
        form = LinkageForm(instance=linkage)
    
    # Pass the linkage object to the template to determine whether it's an edit or add
    return render(request, 'accounts/add_linkage.html', {'form': form, 'linkage': linkage, 'report': linkage.report})

@login_required
def edit_annual_report(request, report_id):
    report = get_object_or_404(AnnualReport, id=report_id)
    if request.method == 'POST':
        form = AnnualReportForm(request.POST, instance=report)
        if form.is_valid():
            form.save()
            return redirect('list_annual_reports')
    else:
        form = AnnualReportForm(instance=report)
    return render(request, 'accounts/create_annual_report.html', {'form': form, 'report': report})
@login_required
def delete_report(request, report_id):
    report = get_object_or_404(AnnualReport, id=report_id)
    report.delete()
    return redirect('list_annual_reports')
@login_required
def add_linkage(request, report_id):
    report = get_object_or_404(AnnualReport, id=report_id)
    if request.method == 'POST':
        form = LinkageForm(request.POST)
        if form.is_valid():
            linkage = form.save(commit=False)
            linkage.report = report
            linkage.save()
            return redirect('view_report', report_id=report.id)
    else:
        form = LinkageForm()
    return render(request, 'accounts/add_linkage.html', {'form': form, 'report': report})
@login_required
def list_annual_reports(request):
    reports = AnnualReport.objects.all().order_by('-year')
    return render(request, 'accounts/list_annual_reports.html', {'reports': reports})
@login_required
def create_quarterly_report(request):
    if request.method == 'POST':
        form = QuarterlyReportForm(request.POST)
        if form.is_valid():
            year = form.cleaned_data['year']
            quarter = form.cleaned_data['quarter']
            existing = QuarterlyReport.objects.filter(year=year, quarter=quarter).first()
            if existing:
                messages.error(request, f"A report for {existing.get_quarter_display()} {year} already exists.")
            else:
                report = form.save()
                messages.success(request, "Quarterly report created successfully.")
                return redirect('view_quarterly_report', report_id=report.id)
    else:
        form = QuarterlyReportForm()
    return render(request, 'accounts/create_quarterly_report.html', {'form': form})
@login_required
def view_quarterly_report(request, report_id):
    report = get_object_or_404(QuarterlyReport, id=report_id)
    projects = report.projects.all()
    return render(request, 'accounts/view_quarterly_report.html', {'report': report, 'projects': projects})
@login_required
def add_extension_project(request, report_id):
    report = get_object_or_404(QuarterlyReport, id=report_id)
    if request.method == 'POST':
        form = ExtensionProjectForm(request.POST)
        if form.is_valid():
            project = form.save(commit=False)
            project.report = report
            project.save()
            return redirect('view_quarterly_report', report_id=report.id)
    else:
        form = ExtensionProjectForm()
    return render(request, 'accounts/add_extension_project.html', {'form': form, 'report': report})

import io
from django.http import FileResponse
from django.shortcuts import get_object_or_404
from reportlab.lib.pagesizes import landscape, letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Image, Paragraph, Spacer, PageBreak
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch
import os
from reportlab.platypus import Table, TableStyle
from networkx import center
from reportlab.lib.units import inch  # Add this at the top if not already imported
from reportlab.platypus import BaseDocTemplate, Frame, PageTemplate
from reportlab.pdfgen.canvas import Canvas

from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import BaseDocTemplate, Frame, PageTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_LEFT
import os
import io
from reportlab.platypus import Paragraph, Table, TableStyle, PageBreak
from reportlab.platypus import BaseDocTemplate, Frame, PageTemplate
from reportlab.lib.pagesizes import landscape, letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_LEFT
from reportlab.lib import colors
from reportlab.pdfgen.canvas import Canvas
from django.http import FileResponse
from django.shortcuts import get_object_or_404
import io
@login_required
def generate_quarterly_pdf(request, report_id):
    report = get_object_or_404(QuarterlyReport, id=report_id)
    activities = report.projects.all()

    buffer = io.BytesIO()

    # Page size and margins
    page_width, page_height = landscape(letter)
    topMargin = 110
    bottomMargin = 70
    leftMargin = 30
    rightMargin = 30

    doc = BaseDocTemplate(
        buffer,
        pagesize=landscape(letter),
        leftMargin=leftMargin,
        rightMargin=rightMargin,
        topMargin=topMargin,
        bottomMargin=bottomMargin,
    )

    frame_height = page_height - topMargin - bottomMargin
    frame_width = page_width - leftMargin - rightMargin
    frame = Frame(leftMargin, bottomMargin, frame_width, frame_height, id='normal')

    doc.addPageTemplates([
        PageTemplate(id='Report', frames=frame, onPage=lambda canvas, doc: draws_header_footer(canvas, doc, report))
    ])

    elements = []
    styles = getSampleStyleSheet()

    wrap_style = ParagraphStyle(
        name='WrapStyle',
        fontSize=8,
        leading=10,
        alignment=TA_LEFT,
    )

    # Table header with spanning
    header = [
        [
            Paragraph("TITLE OF PROJECTS/TRAININGS", wrap_style),
            Paragraph("DATE", wrap_style),
            Paragraph("DURATION<br/>(no. of hrs.)", wrap_style),
            Paragraph("SECTOR", wrap_style),
            Paragraph("BENEFICIARIES", wrap_style), "", "",
            Paragraph("LOCATION/<br/>AREA", wrap_style),
            Paragraph("DEPARTMENT /<br/>PERSONS RESPONSIBLE", wrap_style),
            Paragraph("BUDGET<br/>ALLOCATION", wrap_style),
            Paragraph("REMARKS", wrap_style)
        ],
        ["", "", "", "", "Target No.", "Actual No. M", "Actual No. F", "", "", "", ""]
    ]

    data = []
    for activity in activities:
        data.append([
            Paragraph(activity.title, wrap_style),
            Paragraph(activity.date.strftime('%Y-%m-%d'), wrap_style),
            Paragraph(str(activity.duration_hours), wrap_style),
            Paragraph(activity.sector, wrap_style),
            Paragraph(str(activity.target_no), wrap_style),
            Paragraph(str(activity.actual_male), wrap_style),
            Paragraph(str(activity.actual_female), wrap_style),
            Paragraph(activity.location, wrap_style),
            Paragraph(activity.persons_responsible, wrap_style),
            Paragraph(f"{activity.budget_allocation:,.2f}", wrap_style),
            Paragraph(activity.remarks or "", wrap_style),
        ])

    full_table_data = header + data

    # Table with repeating headers on page breaks
    t = Table(full_table_data, colWidths=[130, 70, 60, 60, 50, 50, 50, 70, 90, 60, 70], repeatRows=2)
    t.setStyle(TableStyle([
        ('SPAN', (4, 0), (6, 0)),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTNAME', (0, 0), (-1, 1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('BACKGROUND', (0, 0), (-1, 1), colors.lightgrey),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ('TOPPADDING', (0, 2), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 2), (-1, -1), 4),
    ]))

    elements.append(t)

    doc.build(elements)
    buffer.seek(0)
    return FileResponse(buffer, as_attachment=False, filename=f"Quarterly_Report_{report.year}.pdf")



from reportlab.lib.pagesizes import landscape, letter
from PIL import Image as PILImage
import os
import tempfile
@login_required
def draws_header_footer(canvas, doc, report):
    canvas.saveState()
    width, height = landscape(letter)

    # Logo settings
    logo_path = os.path.join("static", "image", "logo.png")
    logo_width = 50
    logo_height = 50
    padding = 15  # space between logo and text

    # Approximate text block width (adjust if needed)
    text_block_width = 350

    # Total width of header block (logo + padding + text)
    total_width = logo_width + padding + text_block_width

    # Starting X to center the entire block
    start_x = (width - total_width) / 2

    # Vertical center Y for logo and text
    center_y = height - 50  # adjust to move header up/down

    # Draw logo if exists
    if os.path.exists(logo_path):
        im = PILImage.open(logo_path)
        if im.mode in ('RGBA', 'LA') or (im.mode == 'P' and 'transparency' in im.info):
            bg = PILImage.new("RGB", im.size, (255, 255, 255))
            bg.paste(im, mask=im.split()[-1])
            im = bg
        else:
            im = im.convert("RGB")

        with tempfile.NamedTemporaryFile(delete=False, suffix=".jpg") as tmpfile:
            im.save(tmpfile.name, format='JPEG')
            tmpfile.flush()
            logo_y = center_y - (logo_height / 2)
            canvas.drawImage(tmpfile.name, start_x, logo_y, width=logo_width, height=logo_height, preserveAspectRatio=True)

    # Text starting X (right after logo + padding)
    text_x = start_x + logo_width + padding
    line_height = 15
    text_start_y = center_y + (2 * line_height)

    # Matching font styling from image
    canvas.setFont("Helvetica-Bold", 12)
    canvas.drawCentredString(text_x + (text_block_width / 2), text_start_y, "Republic of the Philippines")

    canvas.setFont("Helvetica-Bold", 14)
    canvas.drawCentredString(text_x + (text_block_width / 2), text_start_y - line_height, "CAVITE STATE UNIVERSITY")

    canvas.setFont("Helvetica", 12)
    canvas.drawCentredString(text_x + (text_block_width / 2), text_start_y - line_height * 2, "Don Severino delas Alas Campus")
    canvas.drawCentredString(text_x + (text_block_width / 2), text_start_y - line_height * 3, "Indang, Cavite")

    canvas.setFont("Helvetica-Bold", 10)
    canvas.drawCentredString(text_x + (text_block_width / 2), text_start_y - line_height * 4,
                             f"{report.get_quarter_display()}, Year {report.year}")

    # Footer
    canvas.setFont("Helvetica", 10)
    canvas.drawString(40, 50, "Prepared by: _____________________  ")
    canvas.drawString(40, 40, "                       Extension Coordinator")
    canvas.drawCentredString(width / 2, 50, "Approved by: _____________________  ")
    canvas.drawCentredString(width / 2, 40, "                Dean/Director")
    canvas.drawRightString(width - 40, 50, "V01-2018-05-30")

    canvas.restoreState()



@login_required
def edit_quarterly_report(request, report_id=None):
    if report_id:
        report = get_object_or_404(QuarterlyReport, id=report_id)
    else:
        report = None

    if request.method == 'POST':
        form = QuarterlyReportForm(request.POST, instance=report)
        if form.is_valid():
            form.save()
            messages.success(request, "Quarterly report updated successfully." if report else "Quarterly report created successfully.")
            return redirect('view_quarterly_report', report_id=report.id)
    else:
        form = QuarterlyReportForm(instance=report)

    return render(request, 'accounts/create_quarterly_report.html', {'form': form, 'report': report})

@login_required
def edit_extension_project(request, project_id):
    project = get_object_or_404(ExtensionProject, id=project_id)
    
    if request.method == 'POST':
        form = ExtensionProjectForm(request.POST, instance=project)
        if form.is_valid():
            form.save()
            messages.success(request, "Extension project updated successfully.")
            return redirect('view_quarterly_report', report_id=project.report.id)
    else:
        form = ExtensionProjectForm(instance=project)
    
    return render(request, 'accounts/add_extension_project.html', {'form': form, 'project': project, 'report': project.report})
@login_required
def delete_quarterly_report(request, report_id):
    report = get_object_or_404(QuarterlyReport, id=report_id)
    report.delete()
    return redirect('create_quarterly_report')
@login_required
def list_quarterly_reports(request):
    reports = QuarterlyReport.objects.all().order_by('-year')
    return render(request, 'accounts/list_quarterly_reports.html', {'reports': reports})


# accounts/views.py
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from .models import Media
from .forms import MediaForm

from django.core.paginator import Paginator
from django.shortcuts import render
from .models import Media, PhotoAlbum, ShowcaseImage

def home2(request):
    # ----------------- Photos (gallery) -----------------
    photos = Media.objects.filter(media_type='photo').order_by('-id')
    photos_per_page = 9
    paginator = Paginator(photos, photos_per_page)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # ----------------- Videos (gallery) -----------------
    videos = Media.objects.filter(media_type='video').order_by('-id')
    total_videos = videos.count()  # <-- IMPORTANT LINE

    videos_per_page = 4
    video_paginator = Paginator(videos, videos_per_page)
    video_page_number = request.GET.get('video_page')
    video_page_obj = video_paginator.get_page(video_page_number)

    # ----------------- Photo Albums -----------------
    albums_list = PhotoAlbum.objects.prefetch_related('photos').all().order_by('-created_at')
    albums_per_page = 9
    album_paginator = Paginator(albums_list, albums_per_page)
    album_page_number = request.GET.get('album_page')
    albums_page = album_paginator.get_page(album_page_number)

    # ----------------- Showcase Images -----------------
    showcase_images = ShowcaseImage.objects.all().order_by('position')

    return render(request, 'accounts/home2.html', {
        'page_obj': page_obj,
        'video_page_obj': video_page_obj,
        'albums': albums_page,
        'showcase_images': showcase_images,
        'total_videos': total_videos,  # <-- FIXED
    })



from django.shortcuts import render, redirect, get_object_or_404
from django.core.paginator import Paginator
from .models import PhotoAlbum, Photo
from .forms import PhotoAlbumForm, PhotoForm

# Album upload
def album_upload(request):
    if request.method == 'POST':
        form = PhotoAlbumForm(request.POST, request.FILES)
        if form.is_valid():
            album = form.save(commit=False)
            album.created_by = request.user
            album.save()
            return redirect('home2')
    else:
        form = PhotoAlbumForm()
    return render(request, 'accounts/album_upload.html', {'form': form})
# accounts/views.py

# accounts/views.py
# views.py
from django.shortcuts import render, get_object_or_404, redirect
from .models import PhotoAlbum, Photo
from .forms import PhotoForm
from django.contrib.auth.decorators import login_required

from django.urls import reverse
from django.shortcuts import redirect, get_object_or_404, render
from django.contrib.auth.decorators import login_required

@login_required
def photo_upload(request, album_id):
    album = get_object_or_404(PhotoAlbum, id=album_id)

    if request.method == 'POST':
        images = request.FILES.getlist('images')  # <-- get all uploaded files
        if images:
            for img in images[:20]:  # Limit to 10 photos
                Photo.objects.create(
                    album=album,
                    image=img,
                    uploaded_by=request.user
                )
            return redirect(f"{reverse('home2')}#album")
    return render(request, 'accounts/photo_upload.html', {
        'album': album
    })


from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from .models import PhotoAlbum, Photo
from .forms import PhotoAlbumForm, PhotoForm


# --- Album Views ---
# --- Album Views ---
@login_required
def edit_album(request, album_id):
    album = get_object_or_404(PhotoAlbum, pk=album_id)
    if request.method == "POST":
        form = PhotoAlbumForm(request.POST, request.FILES, instance=album)
        if form.is_valid():
            form.save()
            messages.success(request, "Album updated successfully.")
            return redirect("home2")
    else:
        form = PhotoAlbumForm(instance=album)
    return render(request, "accounts/album_upload.html", {"form": form, "album": album})



@login_required
def edit_photo(request, photo_id):
    photo = get_object_or_404(Photo, pk=photo_id)
    if request.method == "POST":
        form = PhotoForm(request.POST, request.FILES, instance=photo)
        if form.is_valid():
            form.save()
            messages.success(request, "Photo updated successfully.")
            return redirect("home2")
    else:
        form = PhotoForm(instance=photo)
    return render(request, "accounts/photo_upload.html", {"form": form, "photo": photo, "album": photo.album})


from django.shortcuts import get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from .models import PhotoAlbum, Photo


@login_required
def delete_album(request, album_id):
    album = get_object_or_404(PhotoAlbum, id=album_id)
    if request.method == "POST":
        album.delete()
        return redirect("home2")
    return redirect("home2")

@login_required
def delete_photo(request, photo_id):
    photo = get_object_or_404(Photo, id=photo_id)
    if request.method == "POST":
        photo.delete()
        return redirect("home2")
    return redirect("home2")


def album_photos_api(request, album_id):
    album = PhotoAlbum.objects.get(id=album_id)
    photos = list(album.photos.values('file', 'title', 'description'))
    # prepend MEDIA_URL to file
    for p in photos:
        p['file'] = request.build_absolute_uri(p['file'])
    return JsonResponse(photos, safe=False)



from django.contrib import messages

@login_required
def media_upload(request):
    # Count total videos uploaded (you can also filter by user if needed)
    video_limit = 2
    current_videos = Media.objects.filter(media_type='video').count()   # adjust field name if different

    if request.method == 'POST':
        if current_videos >= video_limit:
            messages.error(request, f"Upload limit reached! You can only upload {video_limit} videos. Delete one first.")
            return redirect('home2')

        form = MediaForm(request.POST, request.FILES)
        if form.is_valid():
            media = form.save(commit=False)
            media.uploaded_by = request.user
            media.save()
            messages.success(request, "Video uploaded successfully!")
            return redirect('home2')

    else:
        form = MediaForm()

    return render(request, 'accounts/media_form.html', {'form': form})


@login_required
def media_edit(request, pk):
    media = get_object_or_404(Media, pk=pk)
    if request.method == 'POST':
        form = MediaForm(request.POST, request.FILES, instance=media)
        if form.is_valid():
            form.save()
            return redirect('home2')
    else:
        form = MediaForm(instance=media)
    return render(request, 'accounts/media_form.html', {'form': form})

@login_required
def media_delete(request, pk):
    media = get_object_or_404(Media, pk=pk)
    if request.method == 'POST':
        media.delete()
        return redirect('home2')
    return render(request, 'accounts/media_confirm_delete.html', {'media': media})


# templates_app/views.py
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from .models import Template
from .forms import TemplateForm
from django.http import HttpResponseForbidden

ALLOWED_ROLES = ['Campus Admin', 'Staff Extensionist']

@login_required
def template_list(request):
    templates = Template.objects.all()
    return render(request, 'accounts/template_list.html', {'templates': templates})


@login_required
def template_upload(request):
    if request.user.account_type not in ALLOWED_ROLES:
        return HttpResponseForbidden("You are not allowed to upload templates.")

    if request.method == 'POST':
        form = TemplateForm(request.POST, request.FILES)
        if form.is_valid():
            template = form.save(commit=False)
            template.uploaded_by = request.user
            template.save()
            return redirect('template_list')
    else:
        form = TemplateForm()
    return render(request, 'accounts/template_form.html', {'form': form, 'action': 'Upload'})


@login_required
def template_edit(request, pk):
    template = get_object_or_404(Template, pk=pk)

    if request.user.account_type not in ALLOWED_ROLES:
        return HttpResponseForbidden("You are not allowed to edit templates.")

    if request.method == 'POST':
        form = TemplateForm(request.POST, request.FILES, instance=template)
        if form.is_valid():
            form.save()
            return redirect('template_list')
    else:
        form = TemplateForm(instance=template)
    return render(request, 'accounts/template_form.html', {'form': form, 'action': 'Edit'})

from django.shortcuts import get_object_or_404, redirect
from django.contrib import messages
from .models import Template  # replace with your actual model name
from django.contrib.auth.decorators import login_required

@login_required
def template_delete(request, pk):
    template = get_object_or_404(Template, pk=pk)

    # Only allow uploader, superuser, or allowed account types to delete
    allowed_account_types = ['Campus Admin', 'Staff Extensionist']
    if request.user == template.uploaded_by or request.user.is_superuser or request.user.account_type in allowed_account_types:
        template.delete()
        messages.success(request, "Template deleted successfully.")
    else:
        messages.error(request, "You do not have permission to delete this template.")

    return redirect('template_list')



from django.db.models.functions import ExtractMonth, ExtractYear
from django.shortcuts import render
from .models import DayTrainingReport
from .utils import QUARTER_CHOICES, get_quarter_for_month

def quarterly_report_list(request):
    reports = DayTrainingReport.objects.annotate(
        month=ExtractMonth("day__date"),
        year=ExtractYear("day__date")
    )

    quarters = set()
    for r in reports:
        if r.month:  # skip if day__date is null
            key, label = get_quarter_for_month(r.month)
            quarters.add((key, r.year, label))

    quarter_order = {key: i for i, (key, _) in enumerate(QUARTER_CHOICES)}
    sorted_quarters = sorted(
        list(quarters),
        key=lambda x: (x[1], quarter_order[x[0]])
    )

    return render(request, "accounts/quarter_list.html", {
        "quarters": sorted_quarters,
    })


from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from .models import DayTrainingReport
from .forms import DayTrainingReportForm
from .utils import QUARTER_CHOICES, get_month_range

# accounts/views.py

from django.shortcuts import get_object_or_404, render, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from .models import DayTrainingReport, DayTrainingEntry, DocumentDay
from .forms import DayTrainingReportForm
from .utils.report_parser import parse_report, DEPARTMENT_NICKNAMES
from .utils import QUARTER_CHOICES, get_month_range


# ---------- HELPER ----------
def nicename_department(dept):
    if not dept:
        return ''
    if hasattr(dept, '__str__'):
        dept = str(dept)
    return DEPARTMENT_NICKNAMES.get(dept, dept)


@login_required
def preview_day_training_reports(request, quarter, year):
    """
    Parse all DayTrainingReports in this quarter/year.
    Create or update DayTrainingEntry objects accordingly.
    """
    start_month, end_month = get_month_range(quarter)

    reports = DayTrainingReport.objects.filter(
        uploaded_at__year=year,
        uploaded_at__month__gte=start_month,
        uploaded_at__month__lte=end_month,
    ).select_related("day", "uploaded_by")

    def safe_float(val):
        try:
            return float(val)
        except (TypeError, ValueError):
            return None

    created, updated = 0, 0

    for rpt in reports:
        existing = DayTrainingEntry.objects.filter(report=rpt).first()
        try:
            parsed = parse_report(rpt.file, uploader_user=rpt.uploaded_by)
        except Exception:
            parsed = None

        if not parsed:
            continue

        # ---------- compute ratings ----------
        r = safe_float(parsed.get("relevance_average"))
        q = safe_float(parsed.get("quality_average"))
        t = safe_float(parsed.get("timeliness_average"))

        total_score = round(r + q + t, 2) if r is not None and q is not None and t is not None else None
        overall_average = round((r + q + t) / 3, 2) if r is not None and q is not None and t is not None else None

        # ---------- department ----------
        raw_dept = parsed.get("department") or getattr(rpt.uploaded_by, "department", "") or ""
        department_value = nicename_department(raw_dept)

        uploader = rpt.uploaded_by
        dept_coordinator = None

        if uploader and uploader.department:
            dept_coordinator = CustomUser.objects.filter(
                account_type="Department Coordinator",
                department=uploader.department
            ).first()

        # ---------- related curricular offering ----------
        def get_related_offering(dept):
            if not dept:
                return ""
            dept = dept.strip().upper()
            mapping = {
                "DEPARTMENT OF COMPUTER STUDIES":
                    "Bachelor of Science in Computer Science, Bachelor of Science in Information Technology",
                "DEPARTMENT OF ARTS AND SCIENCES":
                    "Bachelor of Science in Arts and Sciences",
                "DEPARTMENT OF CRIMINOLOGY":
                    "Bachelor of Science in Criminology",
                "DEPARTMENT OF MANAGEMENT STUDIES (HM)":
                    "Bachelor of Science in Hospitality Management",
                "DEPARTMENT OF MANAGEMENT STUDIES (BM)":
                    "Bachelor of Science in Business Management",
                "DEPARTMENT OF TEACHER EDUCATION":
                    "Bachelor of Science in Secondary Education",
            }
            return mapping.get(dept, "")

        # ----------------- FROZEN FIELDS LOGIC -----------------
        if existing:
            # Freeze old information forever
            coordinator_name = existing.coordinator_name
            coordinator_email = existing.coordinator_email
            contact_person = existing.contact_person
            number_email = existing.number_email
            department_value = existing.department
            related_curricular_offering = getattr(existing, "related_curricular_offering", "")
        else:
            # Default fallback
            coordinator_name = ""
            coordinator_email = ""

            # New upload = assign current coordinator
            if uploader and uploader.account_type == "Extensionist" and dept_coordinator:
                coordinator_name = dept_coordinator.get_full_name()
                coordinator_email = dept_coordinator.email
                contact_person = coordinator_name
                number_email = coordinator_email
            else:
                contact_person = parsed.get("contact_person") or (
                    uploader.get_full_name() if hasattr(uploader, "get_full_name") else getattr(uploader, "username", "")
                )
                number_email = parsed.get("number_email") or getattr(uploader, "email", "")


            # Assign offering only once
            related_curricular_offering = get_related_offering(department_value)

        # ---------- compute totals ----------
        total_participants = parsed.get("total_participants")
        if not total_participants:
            numbers = [
                (parsed.get("male_participants") or 0) + (parsed.get("female_participants") or 0),
                sum([
                    parsed.get("student") or 0,
                    parsed.get("farmer") or 0,
                    parsed.get("fisherfolk") or 0,
                    parsed.get("agricultural_technician") or 0,
                    parsed.get("government_employee") or 0,
                    parsed.get("private_employee") or 0,
                    parsed.get("other_category") or 0
                ])
            ]
            total_participants = next((n for n in numbers if n > 0), 0)

        total_persons_trained = parsed.get("total_persons_trained") or parsed.get("total_by_category") or total_participants or 0

        # ---------- amounts ----------
        amount_charged_to_cvsu = parsed.get("amount_charged_to_cvsu", "0")
        amount_charged_to_partner_agency = parsed.get("amount_charged_to_partner_agency", "0")

        # ---------- common fields ----------
        common_fields = {
            "department": department_value,
            "related_curricular_offering": related_curricular_offering,
            "contact_person": contact_person,
            "number_email": number_email,
            "coordinator_name": coordinator_name,
            "coordinator_email": coordinator_email,
            "project_no": parsed.get("project_no", ""),
            "category": parsed.get("category", ""),
            "title": parsed.get("title") or parsed.get("training_title") or "",
            "date_conducted_text": parsed.get("date_conducted_text", ""),
            "start_date": parsed.get("start_date"),
            "end_date": parsed.get("end_date"),
            "total_by_category": parsed.get("total_by_category"),
            "number_of_days": parsed.get("number_of_days"),
            "male_participants": parsed.get("male_participants"),
            "female_participants": parsed.get("female_participants"),
            "total_participants": total_participants,
            "total_persons_trained": total_persons_trained,
            "student": parsed.get("student"),
            "farmer": parsed.get("farmer"),
            "fisherfolk": parsed.get("fisherfolk"),
            "agricultural_technician": parsed.get("agricultural_technician"),
            "government_employee": parsed.get("government_employee"),
            "private_employee": parsed.get("private_employee"),
            "other_category": parsed.get("other_category"),
            "tvl_solo_parent": parsed.get("tvl_solo_parent"),
            "tvl_4ps": parsed.get("tvl_4ps"),
            "tvl_pwd": parsed.get("tvl_pwd"),
            "tvl_pwd_type": parsed.get("tvl_pwd_type"),
            "total_trainees_surveyed": parsed.get("total_trainees_surveyed"),
            "relevance_counts": parsed.get("relevance_counts") or {},
            "relevance_average": r,
            "quality_counts": parsed.get("quality_counts") or {},
            "quality_average": q,
            "timeliness_counts": parsed.get("timeliness_counts") or {},
            "timeliness_average": t,
            "weight_multiplier": parsed.get("weight_multiplier"),
            "collaborating_agencies": parsed.get("collaborating_agencies", ""),
            "amount_charged_to_cvsu": amount_charged_to_cvsu,
            "amount_charged_to_partner_agency": amount_charged_to_partner_agency,
            "venue": parsed.get("venue", ""),
            "total_score": total_score,
            "overall_average": overall_average,
        }

        if existing:
            # Never overwrite frozen fields
            frozen_fields = [
                "coordinator_name",
                "coordinator_email",
                "contact_person",
                "number_email",
                "department",
                "related_curricular_offering",
            ]

            for field, value in common_fields.items():
                if field in frozen_fields:
                    continue
                setattr(existing, field, value)

            existing.save()
            updated += 1
        else:
            DayTrainingEntry.objects.create(report=rpt, **common_fields)
            created += 1

    return {"created": created, "updated": updated}


@login_required
def quarterly_reports_detail(request, quarter, year):
    """View quarterly reports + parsed training entries."""
    try:
        start_month, end_month = get_month_range(quarter)
    except ValueError:
        messages.error(request, f"Invalid quarter: {quarter}")
        return redirect("quarterly_report_list")

    # Base reports for the quarter
    reports = DayTrainingReport.objects.filter(
        day__date__year=year,
        day__date__month__gte=start_month,
        day__date__month__lte=end_month,
    ).select_related("day__document", "uploaded_by")

    # 🔑 Role-based filtering
    user = request.user
    if user.account_type in [AccountType.DEPARTMENT_COORDINATOR, AccountType.EXTENSIONIST]:
        if user.department:  # filter only if user has department assigned
            reports = reports.filter(day__document__department=user.department)

    # All entries for those reports
    entries = DayTrainingEntry.objects.filter(report__in=reports).order_by("-created_at")

    # Debug info
    parsed_debug = preview_day_training_reports(request, quarter, year)

    # Upload form (only Dept Coordinator can upload)
    form = None
    if user.account_type == AccountType.DEPARTMENT_COORDINATOR:
        if request.method == "POST":
            form = DayTrainingReportForm(request.POST, request.FILES)
            if form.is_valid():
                rpt = form.save(commit=False)
                rpt.uploaded_by = user
                rpt.save()
                messages.success(request, "Day training report uploaded successfully.")
                return redirect("quarterly_reports_detail", quarter=quarter, year=year)
        else:
            form = DayTrainingReportForm()

    context = {
        "quarter": quarter,
        "quarter_label": dict(QUARTER_CHOICES).get(quarter),
        "year": year,
        "reports": reports,
        "entries": entries,
        "form": form,
        "parsed_debug": parsed_debug,
    }
    return render(request, "accounts/quarterly_reports_detail.html", context)



@login_required
def open_day_training_reports(request, quarter, year):

    # ------------------------------------
    # ACCESS CONTROL: Only Campus Admin and Staff Extensionist
    # ------------------------------------
    allowed_roles = ["Campus Admin", "Staff Extensionist"]

    if request.user.account_type not in allowed_roles:
        return redirect('permission_denied')

    # ------------------------------------
    # ORIGINAL FUNCTION LOGIC
    # ------------------------------------
    start_month, end_month = get_month_range(quarter)

    reports = DayTrainingReport.objects.filter(
        day__date__year=year,
        day__date__month__gte=start_month,
        day__date__month__lte=end_month,
    )

    entries = DayTrainingEntry.objects.filter(
        report__in=reports
    ).order_by("department", "-created_at")

    # Auto-fill Related Curricular Offering
    def get_related_offering(dept):
        if not dept:
            return ""

        dept = dept.strip().upper()

        mapping = {
            "DEPARTMENT OF COMPUTER STUDIES":
                "Bachelor of Science in Computer Science, Bachelor of Science in Information Technology",

            "DEPARTMENT OF ARTS AND SCIENCES":
                "Bachelor of Science in Arts and Sciences",

            "DEPARTMENT OF CRIMINOLOGY":
                "Bachelor of Science in Criminology",

            "DEPARTMENT OF MANAGEMENT STUDIES (HM)":
                "Bachelor of Science in Hospitality Management",

            "DEPARTMENT OF MANAGEMENT STUDIES (BM)":
                "Bachelor of Science in Business Management",

            "DEPARTMENT OF TEACHER EDUCATION":
                "Bachelor of Science in Secondary Education",
        }

        return mapping.get(dept, "")

    for e in entries:
        e.related_curricular_offering = get_related_offering(e.department)

    return render(
        request,
        "accounts/open_day_training_reports.html",
        {
            "quarter": quarter,
            "quarter_label": dict(QUARTER_CHOICES).get(quarter),
            "year": year,
            "entries": entries,
        },
    )


     
import io
import io
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import get_column_letter
from django.contrib.auth.decorators import login_required

def col_to_num(col):
    """Convert Excel-style column letters (e.g., A, Z, AA) into a 1-based index."""
    num = 0
    for c in col:
        num = num * 26 + (ord(c.upper()) - 64)
    return num

@login_required
def download_day_training_reports_excel(request, quarter, year):
    import io
    from django.http import HttpResponse
    from openpyxl import Workbook
    from openpyxl.cell.cell import MergedCell
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    
    #  === Get Data ===
    start_month, end_month = get_month_range(quarter)
    reports = DayTrainingReport.objects.filter(
        day__date__year=year,
        day__date__month__gte=start_month,
        day__date__month__lte=end_month,
    )
    entries = DayTrainingEntry.objects.filter(report__in=reports).order_by("department", "-created_at")
    # -------------------------------
    # Auto-fill Related Curricular Offering
    # -------------------------------
    def get_related_offering(dept):
        if not dept:
            return ""
        dept = dept.strip().upper()
        mapping = {
            "DEPARTMENT OF COMPUTER STUDIES":
                "Bachelor of Science in Computer Science, Bachelor of Science in Information Technology",
            "DEPARTMENT OF ARTS AND SCIENCES":
                "Bachelor of Science in Arts and Sciences",
            "DEPARTMENT OF CRIMINOLOGY":
                "Bachelor of Science in Criminology",
            "DEPARTMENT OF MANAGEMENT STUDIES (HM)":
                "Bachelor of Science in Hospitality Management",
            "DEPARTMENT OF MANAGEMENT STUDIES (BM)":
                "Bachelor of Science in Business Management",
            "DEPARTMENT OF TEACHER EDUCATION":
                "Bachelor of Science in Secondary Education",
        }
        return mapping.get(dept, "")
    # Attach to entries
    for e in entries:
        e.related_curricular_offering = get_related_offering(e.department)
    # === Workbook Setup ===
    wb = Workbook()
    ws = wb.active
    ws.title = f"{quarter}_{year}"

    # === Helpers ===
    def safe_set(cell, value):
        """Set value on a cell, handling MergedCell objects gracefully."""
        if isinstance(cell, MergedCell):
            ref = ws.cell(cell.row, cell.column)
            if not isinstance(ref, MergedCell):
                ref.value = value
        else:
            cell.value = value

    # === Styles ===
    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(border_style="thin", color="000000")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)
    header_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    green_fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
    avg_fill = PatternFill(start_color="FABF8F", end_color="FABF8F", fill_type="solid")
    light_green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

    # === MAIN TITLE ===
    ws.merge_cells("A1:BH1")
    safe_set(ws["A1"], f"Cavite State University – Day Training Reports ({quarter.upper()} {year})")
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = center
    ws.row_dimensions[1].height = 30

    # === HEADER STRUCTURE ===
    # Note: inserted new column F -> Related Curricular Offering
    ws.merge_cells("A2:A3"); safe_set(ws["A2"], "Training No.")
    ws.merge_cells("B2:B3"); safe_set(ws["B2"], "Code")
    ws.merge_cells("C2:E2"); safe_set(ws["C2"], "Lead Unit")
    ws.merge_cells("F2:F2"); safe_set(ws["F2"], "Related Curricular Offering")  # NEW
    ws.merge_cells("G2:G3"); safe_set(ws["G2"], "Collaborating Agency/ies")
    ws.merge_cells("H2:H2"); safe_set(ws["H2"], "Contact Person / Training Coordinator")
    ws.merge_cells("I2:I2"); safe_set(ws["I2"], "Project No.")
    ws.merge_cells("J2:J2"); safe_set(ws["J2"], "Category")
    ws.merge_cells("K2:K3"); safe_set(ws["K2"], "Title of Training")
    ws.merge_cells("L2:L3"); safe_set(ws["L2"], "Inclusive Dates")
    ws.merge_cells("M2:M3"); safe_set(ws["M2"], "Venue")
    ws.merge_cells("N2:P2"); safe_set(ws["N2"], "No. of Participants by Sex")  # N,O,P
    ws.merge_cells("Q2:X2"); safe_set(ws["Q2"], "No. of Participants by Category")  # Q..X (8 cols)
    ws.merge_cells("Y2:AB2"); safe_set(ws["Y2"], "For TVL Trainings Only")  # Y,Z,AA,AB (4 cols)
    ws.merge_cells("AC2:AC3"); safe_set(ws["AC2"], "Total Persons Trained")  # AC
    ws.merge_cells("AD2:AH2"); safe_set(ws["AD2"], "Number of Days Trained")  # AD..AH (5 cols)
    ws.merge_cells("AI2:AI3"); safe_set(ws["AI2"], "Number of days trained per weight of training")
    ws.merge_cells("AJ2:AJ3"); safe_set(ws["AJ2"], "Total No. of trainees surveyed")
    ws.merge_cells("AK2:AP2"); safe_set(ws["AK2"], "Client’s Rating (Relevance)")  # AK..AP
    ws.merge_cells("AQ2:AV2"); safe_set(ws["AQ2"], "Client’s Rating (Quality)")    # AQ..AV
    ws.merge_cells("AW2:BB2"); safe_set(ws["AW2"], "Client’s Rating (Timeliness)") # AW..BB
    ws.merge_cells("BC2:BC3"); safe_set(ws["BC2"], "Overall Avg")
    ws.merge_cells("BD2:BD3"); safe_set(ws["BD2"], "Total number of clients requesting trainings")
    ws.merge_cells("BE2:BE3"); safe_set(ws["BE2"], "Total number of requests for trainings responded in the next 3 days")
    ws.merge_cells("BF2:BH2"); safe_set(ws["BF2"], "Estimated Expenses and Source of Fund")  # BF..BH

    # === ROW 3 SUBHEADERS (below headers) ===
    subheaders_row3 = {
        # Lead Unit children
        "C3": "Department / Unit",
        "D3": "Contact Person / Training Coordinator",
        "E3": "Number / Email",
        # NEW column small subtext for Related Curricular Offering
        "F3": "e.g. BS Agriculture",

        # small helper under Contact Person header (kept, shifted)
        "H3": "Training Coordinator",
        "I3": "Use no. indicated under Internally Funded Extension Projects and Externally Funded Projects; indicate NA if not under a project",
        "J3": "TVL–technical, vocational, livelihood; AE–agricultural and environmental trainings; CE–continuing education for professionals; BE–basic education; GAD–Gender and Development; O–others",

        # Participants by sex (N,O,P)
        "N3": "Male", "O3": "Female", "P3": "Total",

        # Participants by category (Q..X)
        "Q3": "Student", "R3": "Farmer", "S3": "Fisherfolk",
        "T3": "Agric. Tech", "U3": "Govt Emp", "V3": "Priv Emp",
        "W3": "Other", "X3": "Total",

        # TVL (Y..AB)
        "Y3": "Solo Parent", "Z3": "4Ps", "AA3": "PWD", "AB3": "Type of Disability",

        # Number of days checkboxes (AD..AH)
        "AD3": "5", "AE3": "3-4", "AF3": "2", "AG3": "1", "AH3": "<8hrs",

        # Relevance (AK..AP)
        "AK3": "5", "AL3": "4", "AM3": "3", "AN3": "2", "AO3": "1", "AP3": "Avg",

        # Quality (AQ..AV)
        "AQ3": "5", "AR3": "4", "AS3": "3", "AT3": "2", "AU3": "1", "AV3": "Avg",

        # Timeliness (AW..BB)
        "AW3": "5", "AX3": "4", "AY3": "3", "AZ3": "2", "BA3": "1", "BB3": "Avg",

        # Estimated expense subheaders
        "BF3": "Amount Charged to CvSU (campus/college/unit)",
        "BG3": "Amount Charged to Partner Agency (PhP)",
        "BH3": "Total Estimated Expense",
    }

    for cell, text in subheaders_row3.items():
        c = ws[cell]
        safe_set(c, text)   # FIX: avoid writing on a MergedCell
        c.font = Font(size=8)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border
        c.fill = header_fill


    # Style first header row (row 2)
    for row_cells in ws.iter_rows(min_row=2, max_row=2):
        for cell in row_cells:
            if cell.value:
                cell.font = bold
                cell.alignment = center
                cell.border = border
                cell.fill = header_fill

    ws.row_dimensions[2].height = 40
    ws.row_dimensions[3].height = 55
    ws.freeze_panes = "A4"

    # === PRECOMPUTE FORMULA COLUMNS (updated for inserted F) ===
    # Total participants (duplicate column) moved from AB -> AC
    col_total_participants = col_to_num("AC")
    # Weight multiplier moved from AH -> AI
    col_multiplier = col_to_num("AI")
    # Total surveyed moved from AI -> AJ
    col_total_surveyed = col_to_num("AJ")

    # === DATA ROWS ===
    row = 4
    color_index = 0
    color_palette = ["FFF2CC", "D9EAD3", "D0E0E3", "F4CCCC", "EAD1DC"]
    last_department = None

    for e in entries:
        days_text = str(e.number_of_days).strip().replace("–", "-").lower()
        if e.department != last_department:
            color_index = (color_index + 1) % len(color_palette)
            fill = PatternFill(start_color=color_palette[color_index],
                               end_color=color_palette[color_index],
                               fill_type="solid")
            last_department = e.department

        # === Row values aligned with new column order (A..BI) ===
        values = [
            row - 3,  # A: Training No.
            "",      # B: Code
            e.department,  # C
            e.contact_person,  # D
            e.number_email,  # E
            e.related_curricular_offering,  # F (NEW)
            e.collaborating_agencies,  # G
            e.contact_person,  # H (Contact Person / Training Coordinator)
            "Internal Funded Extension Projects",  # I
            e.category,  # J
            e.title,  # K
            e.date_conducted_text,  # L
            e.venue,  # M

            # Participants by sex N,O,P
            e.male_participants,  # N
            e.female_participants,  # O
            e.total_participants,  # P

            # Participants by category Q..X
            e.student, e.farmer, e.fisherfolk, e.agricultural_technician,
            e.government_employee, e.private_employee, e.other_category, e.total_by_category,  # Q..X

            # TVL Y..AB
            e.tvl_solo_parent, e.tvl_4ps, e.tvl_pwd, e.tvl_pwd_type,  # Y..AB

            # AC: Total Persons Trained (duplicate of total participants)
            e.total_participants,  # AC

            # Number of days checkboxes AD..AH
            "✓" if days_text == "5" else "",
            "✓" if days_text in ["3", "4"] else "",
            "✓" if days_text == "2" else "",
            "✓" if days_text == "1" else "",
            "✓" if "<8" in days_text else "",  # AD..AH

            # AI: Number of days trained per weight (weight_multiplier)
            e.weight_multiplier,  # AI

            # AJ: Total No. of trainees surveyed
            e.total_trainees_surveyed,  # AJ

            # Relevance AK..AO and Avg at AP
            e.relevance_counts.get("5", 0),
            e.relevance_counts.get("4", 0),
            e.relevance_counts.get("3", 0),
            e.relevance_counts.get("2", 0),
            e.relevance_counts.get("1", 0),
            "",  # AP: Avg (formula set later)

            # Quality AQ..AU and Avg at AV
            e.quality_counts.get("5", 0),
            e.quality_counts.get("4", 0),
            e.quality_counts.get("3", 0),
            e.quality_counts.get("2", 0),
            e.quality_counts.get("1", 0),
            "",  # AV: Avg

            # Timeliness AW..BA and Avg at BB
            e.timeliness_counts.get("5", 0),
            e.timeliness_counts.get("4", 0),
            e.timeliness_counts.get("3", 0),
            e.timeliness_counts.get("2", 0),
            e.timeliness_counts.get("1", 0),
            "",  # BB: Avg

            # BC: Overall Avg (formula later)
            e.overall_average or "",

            # BD..BE
            getattr(e, "total_clients_requesting_trainings", ""),
            getattr(e, "total_requests_responded_next_3_days", ""),

            # Estimated expenses BF..BH
            getattr(e, "amount_charged_to_cvsu", ""),
            getattr(e, "amount_charged_to_partner_agency", ""),
            getattr(e, "estimated_expense_other", ""),

        ]

        # Write row values
        for i, val in enumerate(values, start=1):
            c = ws.cell(row=row, column=i, value=val)
            c.alignment = center
            c.border = border
            c.fill = fill
            if val == "✓":
                c.font = Font(bold=True, color="006100")


        # === Ratings averages formulas (Relevance, Quality, Timeliness) ===
        # Relevance Avg at AP (col AP)
        start_relevance_avg_col = col_to_num("AP")
        start_quality_avg_col = col_to_num("AV")
        start_timeliness_avg_col = col_to_num("BB")

        # For each avg column, build the formula using the five count columns to its left
        for start_col in [start_relevance_avg_col, start_quality_avg_col, start_timeliness_avg_col]:
            # cols for counts are start_col-5 .. start_col-1
            cols_5to1 = [get_column_letter(start_col - 5 + i) + str(row) for i in range(5)]
            # denominator is total surveyed (AJ)
            denom = f"{get_column_letter(col_total_surveyed)}{row}"
            ws.cell(row=row, column=start_col).value = (
                f"=({cols_5to1[0]}*5+{cols_5to1[1]}*4+{cols_5to1[2]}*3+{cols_5to1[3]}*2+{cols_5to1[4]}*1)/{denom}"
            )
            ws.cell(row=row, column=start_col).fill = green_fill
            ws.cell(row=row, column=start_col).alignment = center
            ws.cell(row=row, column=start_col).border = border

        # === Overall Avg (BC) = AVERAGE(AP, AV, BB) ===
        ws.cell(row=row, column=col_to_num("BC")).value = (
            f"=AVERAGE({get_column_letter(col_to_num('AP'))}{row},{get_column_letter(col_to_num('AV'))}{row},{get_column_letter(col_to_num('BB'))}{row})"
        )
        ws.cell(row=row, column=col_to_num("BC")).fill = avg_fill
        ws.cell(row=row, column=col_to_num("BC")).alignment = center
        ws.cell(row=row, column=col_to_num("BC")).border = border

        ws.row_dimensions[row].height = 70
        row += 1

    # === Column Widths (with new F inserted) ===
    widths = {
        "A":12,"B":20,"C":25,"D":25,"E":20,"F":25,"G":35,"H":20,"I":25,"J":35,"K":20,"L":20,
        "M":10,"N":10,"O":10,"P":10,"Q":10,"R":10,"S":10,"T":10,"U":15,"V":15,"W":15,"X":20,
        "Y":15,"Z":8,"AA":15,"AB":15,"AC":15,"AD":10,"AE":12,"AF":15,"AG":15,"AH":8,"AI":15,
        "AJ":15,"AK":8,"AL":10,"AM":8,"AN":8,"AO":8,"AP":8,"AQ":8,"AR":10,"AS":8,"AT":8,"AU":8,
        "AV":8,"AW":8,"AX":10,"AY":15,"AZ":20,"BA":18,"BB":18,"BC":20,"BD":18,"BE":18,"BF":12,
        "BG":12,"BH":12
    }
    for k, v in widths.items():
        ws.column_dimensions[k].width = v

    # === Thin border for entire used range ===
    thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row_cells in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row_cells:
            # Avoid overwriting merged cell style in some excel versions (safe)
            try:
                cell.border = thin_border
            except Exception:
                pass

    # === OUTPUT ===
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    response = HttpResponse(
        output,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="DayTrainingReports_{quarter}_{year}.xlsx"'
    return response


from django.shortcuts import redirect, get_object_or_404
from django.http import JsonResponse
from .models import ShowcaseImage

from django.shortcuts import redirect, render
from .models import ShowcaseImage

def upload_image(request):
    if request.method == "POST" and request.FILES.get("image"):
        # Get max current position
        last = ShowcaseImage.objects.all().order_by('-position').first()
        pos = last.position + 1 if last else 0
        ShowcaseImage.objects.create(image=request.FILES["image"], position=pos)
        return redirect("home2")
    return redirect("home2")


def delete_image(request, image_id):
    img = get_object_or_404(ShowcaseImage, id=image_id)
    img.delete()
    return redirect(request.META.get("HTTP_REFERER", "/"))


def reorder_images(request):
    if request.method == "POST":
        order_list = request.POST.getlist("order[]")
        for index, img_id in enumerate(order_list):
            ShowcaseImage.objects.filter(id=img_id).update(position=index)
        return JsonResponse({"status": "success"})
    return JsonResponse({"status": "failed"}, status=400)

from django.shortcuts import get_object_or_404, redirect
from .models import ShowcaseImage

def move_showcase_image(request, image_id, direction):
    if request.method == "POST":
        img = get_object_or_404(ShowcaseImage, id=image_id)
        if direction == "up":
            prev_img = ShowcaseImage.objects.filter(position__lt=img.position).order_by('-position').first()
            if prev_img:
                # Swap positions
                img.position, prev_img.position = prev_img.position, img.position
                img.save()
                prev_img.save()
        elif direction == "down":
            next_img = ShowcaseImage.objects.filter(position__gt=img.position).order_by('position').first()
            if next_img:
                # Swap positions
                img.position, next_img.position = next_img.position, img.position
                img.save()
                next_img.save()
    return redirect(request.META.get("HTTP_REFERER", "/"))

from django.shortcuts import get_object_or_404, redirect
from .models import ShowcaseImage

def edit_showcase_image(request, image_id):
    if request.method == "POST" and request.FILES.get("image"):
        img = get_object_or_404(ShowcaseImage, id=image_id)
        img.image = request.FILES["image"]  # replace image
        img.save()  # keep the same position
    return redirect(request.META.get("HTTP_REFERER", "/"))
